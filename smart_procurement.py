# -*- coding: utf-8 -*-
"""
 스마트 발주 시스템 (Smart Procurement System)
Streamlit 기반 자동화 발주 솔루션
"""

import streamlit as st
import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import os
import numpy as np
import json

# Claude API (설치: pip install anthropic)
try:
    import anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False

# 페이지 설정
st.set_page_config(
    page_title=" 스마트 발주 시스템",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 발주 기록 함수
def record_order_to_excel(psi_file_path, order_data):
    """
    발주 내역을 PSI 파일의 '발주리스트' 시트에 기록

    Parameters:
    - psi_file_path: PSI 엑셀 파일 경로
    - order_data: dict with keys: SKU코드, 제품명, ABC등급, XYZ등급, 현재고, 발주량,
                  매입원가, 재고소진일, 리드타임
    """
    try:
        # 파일 열기
        wb = openpyxl.load_workbook(psi_file_path)

        # 발주리스트 시트 확인 및 자동 생성
        if '발주리스트' not in wb.sheetnames:
            # 시트가 없으면 자동 생성
            ws = wb.create_sheet('발주리스트')

            # 헤더 추가
            from openpyxl.styles import Font, PatternFill, Alignment
            headers = ['발주일', 'SKU#', '제품명', 'ABC/XYZ', '현재고', '발주량',
                      '구매원가', '발주 전 재고소진일', '발주 후 재고소진일', '예상입고일']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(1, col_idx, header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # 헤더 폰트를 흰색으로
                cell.font = Font(bold=True, size=11, color="FFFFFF")

            # 컬럼 너비 조정
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 18
            ws.column_dimensions['I'].width = 18
            ws.column_dimensions['J'].width = 18

            st.info("✅ '발주리스트' 시트를 자동으로 생성했습니다!")
        else:
            ws = wb['발주리스트']

        # 다음 빈 행 찾기 (2행부터 데이터 시작)
        next_row = 2
        while ws.cell(next_row, 1).value is not None:
            next_row += 1

        # 발주 데이터 작성
        발주일 = datetime.now().strftime('%Y-%m-%d')  # 날짜만 (시간 제외)
        리드타임 = int(order_data.get('리드타임', 30))  # int로 변환
        예상입고일 = (datetime.now() + timedelta(days=리드타임)).strftime('%Y-%m-%d')  # 날짜만

        # 발주 후 재고소진일 계산
        발주후재고 = float(order_data['현재고']) + float(order_data['발주량'])
        일평균판매 = float(order_data.get('일평균판매', 0))
        if 일평균판매 > 0:
            발주후재고소진일 = 발주후재고 / 일평균판매
        else:
            발주후재고소진일 = 999

        # A: 발주일 (날짜만)
        ws.cell(next_row, 1, 발주일)

        # B: SKU#
        ws.cell(next_row, 2, order_data['SKU코드'])

        # C: 제품명
        ws.cell(next_row, 3, order_data['제품명'])

        # D: ABC/XYZ
        abc_xyz = f"{order_data.get('ABC등급', 'N/A')}/{order_data.get('XYZ등급', 'N/A')}"
        ws.cell(next_row, 4, abc_xyz)

        # E: 현재고
        ws.cell(next_row, 5, order_data['현재고'])

        # F: 발주량
        ws.cell(next_row, 6, order_data['발주량'])

        # G: 구매원가
        ws.cell(next_row, 7, order_data.get('매입원가', 0))

        # H: 발주 전 재고소진일
        ws.cell(next_row, 8, order_data.get('재고소진일', 0))

        # I: 발주 후 재고소진일
        ws.cell(next_row, 9, 발주후재고소진일)

        # J: 예상입고일
        ws.cell(next_row, 10, 예상입고일)

        # 저장
        wb.save(psi_file_path)
        wb.close()

        return True

    except Exception as e:
        st.error(f"❌ 발주 기록 중 오류: {str(e)}")
        return False

# ==================== AI 수요 예측 엔진 ====================

def forecast_demand_simple(daily_sales, forecast_days=30):
    """
    단순 수요 예측 (이동평균 + 지수평활법)

    Parameters:
    - daily_sales: 일평균 판매량
    - forecast_days: 예측 기간 (일)

    Returns:
    - dict: 예측 결과 및 신뢰도
    """
    if daily_sales <= 0:
        return {
            'forecast_daily': 0,
            'forecast_total': 0,
            'confidence': 0,
            'trend': 'unknown',
            'method': 'zero_sales'
        }

    # 단순 예측: 현재 판매량 유지 가정
    forecast_daily = daily_sales
    forecast_total = daily_sales * forecast_days

    return {
        'forecast_daily': round(forecast_daily, 1),
        'forecast_total': round(forecast_total, 0),
        'confidence': 70,  # 기본 신뢰도
        'trend': 'stable',
        'method': 'simple_average'
    }

def forecast_demand_advanced(sales_history, forecast_days=30):
    """
    고급 수요 예측 (트렌드 분석 포함)

    Parameters:
    - sales_history: 과거 판매 데이터 리스트 (최근 순)
    - forecast_days: 예측 기간 (일)

    Returns:
    - dict: 예측 결과, 트렌드, 신뢰도
    """
    if not sales_history or len(sales_history) == 0:
        return forecast_demand_simple(0, forecast_days)

    # numpy 배열로 변환
    sales = np.array(sales_history)

    # 기본 통계
    mean_sales = np.mean(sales)
    std_sales = np.std(sales)

    # 트렌드 분석 (선형 회귀)
    if len(sales) >= 3:
        x = np.arange(len(sales))
        coefficients = np.polyfit(x, sales, 1)  # 1차 선형 회귀
        trend_slope = coefficients[0]

        # 트렌드 판단
        if trend_slope > mean_sales * 0.05:  # 5% 이상 증가
            trend = 'increasing'
            trend_factor = 1.1  # 10% 증가 반영
        elif trend_slope < -mean_sales * 0.05:  # 5% 이상 감소
            trend = 'decreasing'
            trend_factor = 0.9  # 10% 감소 반영
        else:
            trend = 'stable'
            trend_factor = 1.0
    else:
        trend = 'insufficient_data'
        trend_factor = 1.0

    # 지수평활법 (Exponential Smoothing)
    alpha = 0.3  # 평활 계수
    if len(sales) >= 2:
        forecast_daily = sales[-1] * alpha + (1 - alpha) * mean_sales
    else:
        forecast_daily = mean_sales

    # 트렌드 반영
    forecast_daily *= trend_factor
    forecast_total = forecast_daily * forecast_days

    # 신뢰도 계산 (변동 계수 기반)
    cv = (std_sales / mean_sales) if mean_sales > 0 else 1.0
    if cv < 0.2:
        confidence = 90
    elif cv < 0.5:
        confidence = 75
    elif cv < 1.0:
        confidence = 60
    else:
        confidence = 40

    return {
        'forecast_daily': round(forecast_daily, 1),
        'forecast_total': round(forecast_total, 0),
        'confidence': confidence,
        'trend': trend,
        'trend_slope': round(trend_slope, 1) if 'trend_slope' in locals() else 0,
        'method': 'exponential_smoothing',
        'cv': round(cv, 1)
    }

# ==================== 동적 발주점 계산 ====================

def calculate_dynamic_reorder_point(row, forecast_data=None, safety_factor=1.5):
    """
    동적 발주점 계산 (수요 예측 기반)

    Parameters:
    - row: 품목 데이터 (pandas Series)
    - forecast_data: 수요 예측 결과 (dict)
    - safety_factor: 안전계수 (ABC 등급별)

    Returns:
    - dict: 발주점, 안전재고, 권장발주량
    """
    # 기본값
    일평균판매 = row.get('일평균판매', 0)
    리드타임 = row.get('리드타임', 30)
    현재고 = row.get('현재고', 0)
    ABC등급 = row.get('ABC등급', 'C')

    # ABC 등급별 안전계수
    safety_factors = {'A': 1.5, 'B': 1.2, 'C': 1.0}
    safety_factor = safety_factors.get(ABC등급, 1.2)

    # 수요 예측 데이터 사용 (있으면)
    if forecast_data:
        예측_일평균판매 = forecast_data['forecast_daily']
        신뢰도 = forecast_data['confidence'] / 100

        # 신뢰도가 낮으면 안전계수 증가
        if 신뢰도 < 0.6:
            safety_factor *= 1.2
    else:
        예측_일평균판매 = 일평균판매

    # 동적 안전재고 = 예측 판매량 × 리드타임 × 안전계수
    안전재고 = 예측_일평균판매 * 리드타임 * safety_factor

    # 동적 발주점 = 안전재고 + (예측 판매량 × 리드타임)
    발주점 = 안전재고 + (예측_일평균판매 * 리드타임)

    # 권장발주량 계산
    if 현재고 < 발주점:
        목표재고 = 발주점 * 1.5  # 발주점의 1.5배를 목표로
        권장발주량 = max(0, 목표재고 - 현재고)

        # MOQ 반영 (None 처리)
        MOQ = row.get('MOQ') or 0
        if MOQ > 0 and 권장발주량 > 0:
            # MOQ의 배수로 올림
            권장발주량 = np.ceil(권장발주량 / MOQ) * MOQ
    else:
        권장발주량 = 0

    return {
        '안전재고': round(안전재고, 0),
        '발주점': round(발주점, 0),
        '권장발주량': round(권장발주량, 0),
        '예측_일평균판매': round(예측_일평균판매, 1),
        '현재_재고소진일': round(현재고 / 예측_일평균판매, 1) if 예측_일평균판매 > 0 else 999,
        '발주필요': 현재고 < 발주점
    }

# ==================== 자동 발주 생성 ====================

def generate_auto_orders(df_analysis):
    """
    자동 발주 생성 (승인 대기)

    Parameters:
    - df_analysis: 품목 분석 데이터프레임

    Returns:
    - list: 자동 발주 추천 목록
    """
    auto_orders = []

    for idx, row in df_analysis.iterrows():
        # 발주 필요 여부 확인
        if not row.get('발주필요', False):
            continue

        # 수요 예측
        일평균판매 = row.get('일평균판매', 0)
        forecast = forecast_demand_simple(일평균판매, 30)

        # 동적 발주점 계산
        reorder_info = calculate_dynamic_reorder_point(row, forecast)

        # 우선순위 계산
        ABC등급 = row.get('ABC등급', 'C')
        재고상태 = row.get('재고상태', '')

        if '부족' in 재고상태:
            priority = 'HIGH'
            priority_score = 10
        elif '재주문' in 재고상태:
            priority = 'MEDIUM'
            priority_score = 5
        else:
            priority = 'LOW'
            priority_score = 1

        # ABC 등급 추가 점수
        if ABC등급 == 'A':
            priority_score += 5
        elif ABC등급 == 'B':
            priority_score += 2

        # 자동 발주 정보 생성
        order = {
            'SKU코드': row.get('SKU코드', ''),
            '제품명': row.get('제품명', ''),
            'ABC등급': ABC등급,
            'XYZ등급': row.get('XYZ등급', ''),
            '현재고': row.get('현재고', 0),
            '안전재고': reorder_info['안전재고'],
            '발주점': reorder_info['발주점'],
            '권장발주량': reorder_info['권장발주량'],
            'MOQ': row.get('MOQ', 0),
            '리드타임': row.get('리드타임', 30),
            '예측_일평균판매': forecast['forecast_daily'],
            '예측_신뢰도': forecast['confidence'],
            '예측_트렌드': forecast['trend'],
            '우선순위': priority,
            '우선순위_점수': priority_score,
            '재고상태': 재고상태,
            '재고소진일': reorder_info['현재_재고소진일'],
            '매입원가': row.get('매입원가', 0),
            '예상_발주금액': reorder_info['권장발주량'] * row.get('매입원가', 0)
        }

        auto_orders.append(order)

    # 우선순위 점수로 정렬
    auto_orders.sort(key=lambda x: x['우선순위_점수'], reverse=True)

    return auto_orders

# 모던 블루 스타일 CSS (Deepflow Style)
st.markdown("""
<style>
    /* 최우선 강제 라이트 모드 */
    :root {
        color-scheme: light !important;
    }

    html, body {
        background-color: #F8F9FA !important;
        color: #1E293B !important;
    }

    /* 전체 배경 - 밝은 회색 */
    .stApp {
        background-color: #F8F9FA !important;
        overflow-y: auto !important;
        color: #1E293B !important;
    }

    .main {
        background-color: #F8F9FA !important;
        overflow-y: auto !important;
        color: #1E293B !important;
    }

    /* 모든 요소 기본 색상 */
    * {
        color-scheme: light !important;
    }

    section[data-testid="stSidebar"] > div {
        background-color: #F5F5F5 !important;
    }

    [data-testid="stAppViewContainer"] {
        background-color: #F8F9FA !important;
    }

    [data-testid="stHeader"] {
        background-color: #F8F9FA !important;
    }

    /* 메인 컨테이너 */
    .block-container {
        padding-top: 1rem !important;
        padding-bottom: 2rem !important;
        max-width: 100% !important;
    }

    /* 상단 여백 제거 */
    .main .block-container {
        margin-top: 0 !important;
        padding-top: 1rem !important;
    }

    /* 사이드바 - 연한 보라/회색 - 강제 라이트 모드 */
    [data-testid="stSidebar"],
    section[data-testid="stSidebar"],
    .css-1d391kg,
    .css-1lcbmhc {
        background-color: #F5F5F5 !important;
    }

    [data-testid="stSidebar"] *,
    section[data-testid="stSidebar"] *,
    .css-1d391kg *,
    .css-1lcbmhc * {
        color: #1E293B !important;
    }

    /* 사이드바 내부 모든 컴포넌트 배경 */
    [data-testid="stSidebar"] div,
    [data-testid="stSidebar"] section,
    [data-testid="stSidebar"] [data-testid="stExpander"] {
        background-color: transparent !important;
    }

    /* 사이드바 위젯 라벨 */
    [data-testid="stSidebar"] label {
        color: #1E293B !important;
    }

    /* 메트릭 카드 - 연한 민트 */
    [data-testid="stMetric"] {
        background: linear-gradient(135deg, #C8E6C9 0%, #E0F2E9 100%);
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 4px 12px rgba(200, 230, 201, 0.2);
        color: #2C3E50 !important;
    }

    [data-testid="stMetric"] * {
        color: #2C3E50 !important;
    }

    [data-testid="stMetricLabel"] {
        color: #5A6C7D !important;
        font-size: 0.9rem !important;
    }

    [data-testid="stMetricValue"] {
        color: #2C3E50 !important;
        font-size: 2rem !important;
        font-weight: 700 !important;
    }

    /* 탭 스타일 - 명확한 테두리와 배경 */
    .stTabs [data-baseweb="tab-list"] {
        gap: 12px !important;
        background-color: #F8F9FA !important;
        border-radius: 12px !important;
        padding: 8px !important;
        border: 2px solid #E5E7EB !important;
    }

    .stTabs [data-baseweb="tab"] {
        height: 50px !important;
        border-radius: 8px !important;
        color: #1E293B !important;
        font-weight: 600 !important;
        background-color: white !important;
        border: 2px solid #E5E7EB !important;
        padding: 8px 24px !important;
        transition: all 0.3s ease !important;
    }

    .stTabs [data-baseweb="tab"]:hover {
        background-color: #F8F9FA !important;
        border-color: #81C784 !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 4px 8px rgba(0,0,0,0.1) !important;
    }

    .stTabs [aria-selected="true"] {
        background-color: #81C784 !important;
        color: white !important;
        font-weight: 700 !important;
        border: 2px solid #81C784 !important;
        box-shadow: 0 4px 12px rgba(129,199,132,0.4) !important;
    }

    /* 버튼 - Primary */
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #C8E6C9 0%, #E0F2E9 100%);
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        box-shadow: 0 4px 12px rgba(200, 230, 201, 0.3);
        transition: all 0.3s ease;
    }

    .stButton > button[kind="primary"]:hover {
        box-shadow: 0 6px 16px rgba(68, 97, 242, 0.4);
        transform: translateY(-2px);
    }

    /* 버튼 - Secondary */
    .stButton > button {
        background-color: white;
        color: #C8E6C9;
        border: 2px solid #C8E6C9;
        border-radius: 8px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        transition: all 0.3s ease;
    }

    .stButton > button:hover {
        background-color: #E8E8E8;
    }

    /* 알림 박스 - 긴급 (빨강) */
    .stAlert[data-baseweb="notification"][kind="error"] {
        background-color: #FEE2E2;
        color: #991B1B;
        border-left: 4px solid #EF4444;
        border-radius: 8px;
        padding: 1rem;
    }

    /* 알림 박스 - 경고 (노랑) */
    .stAlert[data-baseweb="notification"][kind="warning"] {
        background-color: #FEF3C7;
        color: #92400E;
        border-left: 4px solid #F59E0B;
        border-radius: 8px;
        padding: 1rem;
    }

    /* 알림 박스 - 성공 (초록) */
    .stAlert[data-baseweb="notification"][kind="success"] {
        background-color: #D1FAE5;
        color: #065F46;
        border-left: 4px solid #10B981;
        border-radius: 8px;
        padding: 1rem;
    }

    /* 알림 박스 - 정보 (파랑) */
    .stAlert[data-baseweb="notification"][kind="info"] {
        background-color: #DBEAFE;
        color: #1E40AF;
        border-left: 4px solid #C8E6C9;
        border-radius: 8px;
        padding: 1rem;
    }

    /* 데이터 테이블 */
    .dataframe {
        background-color: white !important;
        border-radius: 12px;
        border: 1px solid #E5E7EB;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.05);
    }

    .dataframe thead tr th {
        background-color: #F8F9FA !important;
        color: #1E293B !important;
        font-weight: 700 !important;
        padding: 1rem !important;
        border-bottom: 2px solid #E5E7EB !important;
    }

    .dataframe tbody tr td {
        color: #475569 !important;
        padding: 0.875rem !important;
        border-bottom: 1px solid #F1F5F9 !important;
    }

    .dataframe tbody tr:hover {
        background-color: #F8F9FA !important;
    }

    /* Streamlit 데이터프레임 - 강제 라이트 모드 */
    [data-testid="stDataFrame"],
    [data-testid="stDataFrame"] > div,
    [data-testid="stDataFrame"] div[data-testid="stDataFrameResizable"] {
        background-color: white !important;
        border-radius: 12px;
        overflow: hidden;
    }

    [data-testid="stDataFrame"] table {
        border-collapse: separate !important;
        border-spacing: 0 !important;
        background-color: white !important;
    }

    [data-testid="stDataFrame"] thead th {
        background-color: #F8F9FA !important;
        color: #1E293B !important;
        font-weight: 700 !important;
        text-align: left !important;
        padding: 1rem !important;
        border-bottom: 2px solid #E5E7EB !important;
    }

    [data-testid="stDataFrame"] tbody tr {
        background-color: white !important;
    }

    [data-testid="stDataFrame"] tbody td {
        background-color: white !important;
        color: #1E293B !important;
        padding: 0.875rem !important;
        border-bottom: 1px solid #F1F5F9 !important;
    }

    [data-testid="stDataFrame"] tbody tr:hover {
        background-color: #F8F9FA !important;
    }

    [data-testid="stDataFrame"] tbody tr:hover td {
        background-color: #F8F9FA !important;
    }

    /* 데이터 그리드 내부 모든 요소 */
    [data-testid="stDataFrame"] * {
        color: #1E293B !important;
    }

    /* 체크박스나 다른 요소들도 밝게 */
    [data-testid="stDataFrame"] [role="gridcell"] {
        background-color: white !important;
    }

    /* 입력 필드 - 강제 라이트 모드 */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stSelectbox > div > div,
    .stMultiselect > div > div,
    [data-baseweb="select"],
    [data-baseweb="input"] {
        border: 2px solid #E5E7EB !important;
        border-radius: 8px !important;
        padding: 0.75rem !important;
        background-color: white !important;
        color: #1E293B !important;
        transition: all 0.3s ease !important;
    }

    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus,
    .stSelectbox > div > div:focus-within,
    .stMultiselect > div > div:focus-within {
        border-color: #C8E6C9 !important;
        box-shadow: 0 0 0 3px rgba(68, 97, 242, 0.1) !important;
    }

    /* Multiselect 태그 - 명확한 테두리 */
    [data-baseweb="tag"] {
        background-color: #81C784 !important;
        color: white !important;
        border: 3px solid #4CAF50 !important;
        border-radius: 8px !important;
        padding: 6px 12px !important;
        font-weight: 600 !important;
        box-shadow: 0 2px 6px rgba(76, 175, 80, 0.3) !important;
        margin: 4px !important;
    }

    /* 태그 호버 효과 */
    [data-baseweb="tag"]:hover {
        background-color: #66BB6A !important;
        border-color: #388E3C !important;
        box-shadow: 0 4px 8px rgba(76, 175, 80, 0.4) !important;
        transform: translateY(-1px) !important;
    }

    /* 태그 내부 텍스트 */
    [data-baseweb="tag"] span {
        color: white !important;
        font-weight: 600 !important;
    }

    /* 태그 X 버튼 */
    [data-baseweb="tag"] svg {
        color: white !important;
        opacity: 0.9 !important;
    }

    [data-baseweb="tag"] svg:hover {
        opacity: 1 !important;
    }

    /* Multiselect 드롭다운 */
    [data-baseweb="popover"] {
        background-color: white !important;
    }

    [role="listbox"] {
        background-color: white !important;
    }

    [role="option"] {
        background-color: white !important;
        color: #1E293B !important;
    }

    [role="option"]:hover {
        background-color: #F8F9FA !important;
    }

    /* 슬라이더 */
    .stSlider > div > div > div {
        background-color: #E5E7EB;
    }

    .stSlider > div > div > div > div {
        background-color: #C8E6C9;
    }

    /* 체크박스 */
    .stCheckbox > label > div {
        background-color: white;
        border: 2px solid #E5E7EB;
        border-radius: 4px;
    }

    .stCheckbox > label > div[data-checked="true"] {
        background-color: #C8E6C9;
        border-color: #C8E6C9;
    }

    /* 라디오 버튼 */
    .stRadio > label > div {
        background-color: white;
        border: 2px solid #E5E7EB;
        border-radius: 50%;
    }

    .stRadio > label > div[data-checked="true"] {
        background-color: #C8E6C9;
        border-color: #C8E6C9;
    }

    /* Expander - 강제 라이트 모드 */
    .streamlit-expanderHeader,
    [data-testid="stExpander"] summary,
    [data-testid="stExpander"] > details > summary {
        background-color: white !important;
        border: 1px solid #E5E7EB !important;
        border-radius: 8px !important;
        color: #1E293B !important;
        font-weight: 600 !important;
    }

    .streamlit-expanderHeader:hover,
    [data-testid="stExpander"] summary:hover {
        background-color: #F8F9FA !important;
    }

    .streamlit-expanderContent,
    [data-testid="stExpander"] > details > div,
    [data-testid="stExpanderDetails"] {
        background-color: white !important;
        border: 1px solid #E5E7EB !important;
        border-top: none !important;
        border-radius: 0 0 8px 8px !important;
    }

    /* Expander 내부 모든 요소 */
    [data-testid="stExpander"] *,
    .streamlit-expanderHeader *,
    .streamlit-expanderContent * {
        color: #1E293B !important;
    }

    /* 차트 */
    .js-plotly-plot {
        background-color: white !important;
        border-radius: 12px;
        padding: 1rem;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.05);
    }

    /* 헤더 */
    h1, h2, h3, h4, h5, h6 {
        color: #1E293B !important;
        font-weight: 700 !important;
    }

    /* 텍스트 - 더 진한 색상 */
    p, span, div {
        color: #1E293B !important;
    }

    /* 마크다운 */
    .stMarkdown {
        color: #1E293B !important;
    }

    /* 위젯 레이블 - 명확한 색상 */
    label, .stLabel, [data-testid="stWidgetLabel"] {
        color: #0F172A !important;
        font-weight: 600 !important;
    }

    /* 입력 필드 레이블 */
    .stTextInput label,
    .stNumberInput label,
    .stSelectbox label,
    .stMultiselect label,
    .stSlider label,
    .stRadio label,
    .stCheckbox label {
        color: #0F172A !important;
        font-weight: 600 !important;
        font-size: 0.95rem !important;
    }

    /* 섹션 제목 */
    .stSubheader {
        color: #0F172A !important;
        font-weight: 700 !important;
    }

    /* 다운로드 버튼 */
    .stDownloadButton > button {
        background-color: white;
        color: #C8E6C9;
        border: 2px solid #C8E6C9;
        border-radius: 8px;
        padding: 0.75rem 2rem;
        font-weight: 600;
    }

    .stDownloadButton > button:hover {
        background-color: #E8E8E8;
    }

    /* 프로그레스 바 */
    .stProgress > div > div > div {
        background-color: #C8E6C9;
    }

    /* 스피너 */
    .stSpinner > div {
        border-top-color: #C8E6C9 !important;
    }

    /* 구분선 */
    hr {
        border: none;
        border-top: 2px solid #E5E7EB;
        margin: 2rem 0;
    }

    /* 코드 블록 */
    .stCodeBlock {
        background-color: #F8F9FA;
        border: 1px solid #E5E7EB;
        border-radius: 8px;
    }

    code {
        background-color: #F8F9FA;
        color: #C8E6C9;
        padding: 0.2rem 0.4rem;
        border-radius: 4px;
        font-family: 'Monaco', 'Menlo', monospace;
    }

    /* 카드 스타일 (커스텀) */
    .card {
        background-color: white;
        border-radius: 12px;
        padding: 1.5rem;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.05);
        border: 1px solid #E5E7EB;
        margin-bottom: 1rem;
    }

    .card-header {
        font-size: 1.25rem;
        font-weight: 700;
        color: #1E293B;
        margin-bottom: 1rem;
    }

    .card-body {
        color: #475569;
    }

    /* 배지 스타일 */
    .badge {
        display: inline-block;
        padding: 0.25rem 0.75rem;
        border-radius: 12px;
        font-size: 0.875rem;
        font-weight: 600;
    }

    .badge-primary {
        background-color: #DBEAFE;
        color: #1E40AF;
    }

    .badge-success {
        background-color: #D1FAE5;
        color: #065F46;
    }

    .badge-warning {
        background-color: #FEF3C7;
        color: #92400E;
    }

    .badge-danger {
        background-color: #FEE2E2;
        color: #991B1B;
    }

    /* 반응형 */
    @media (max-width: 768px) {
        .block-container {
            padding-left: 1rem !important;
            padding-right: 1rem !important;
        }

        [data-testid="stMetric"] {
            padding: 1rem;
        }
    }

    /* 애니메이션 */
    @keyframes fadeIn {
        from {
            opacity: 0;
            transform: translateY(10px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }

    .stApp > * {
        animation: fadeIn 0.3s ease-in-out;
    }

    /* 호버 효과 */
    .card:hover {
        box-shadow: 0 4px 16px rgba(68, 97, 242, 0.15);
        transform: translateY(-2px);
        transition: all 0.3s ease;
    }

    /* 링크 */
    a {
        color: #C8E6C9 !important;
        text-decoration: none;
        font-weight: 600;
    }

    a:hover {
        text-decoration: underline;
    }

    /* Form 스타일 */
    .stForm {
        background-color: white;
        border-radius: 12px;
        padding: 1.5rem;
        border: 1px solid #E5E7EB;
    }

    /* 파일 업로더 - 강제 라이트 모드 */
    .stFileUploader,
    [data-testid="stFileUploader"],
    section[data-testid="stFileUploadDropzone"] {
        background: linear-gradient(135deg, #FFFFFF 0%, #F8F9FA 100%) !important;
        border: 3px dashed #C8E6C9 !important;
        border-radius: 16px !important;
        padding: 3rem 2rem !important;
        transition: all 0.3s ease !important;
    }

    .stFileUploader:hover,
    [data-testid="stFileUploader"]:hover {
        border-color: #7BA591 !important;
        background: linear-gradient(135deg, #F8F9FA 0%, #F5F5F5 100%) !important;
        box-shadow: 0 4px 12px rgba(200, 230, 201, 0.3) !important;
        transform: scale(1.01) !important;
    }

    .stFileUploader label,
    .stFileUploader *,
    [data-testid="stFileUploader"] *,
    [data-testid="stFileUploader"] label {
        color: #2C3E50 !important;
        font-size: 1.2rem !important;
        font-weight: 700 !important;
        text-align: center !important;
    }

    /* 드래그 앤 드롭 영역 */
    [data-testid="stFileUploaderDropzone"],
    [data-testid="stFileUploadDropzone"],
    section[data-testid="stFileUploadDropzone"] > div {
        background-color: #FFFFFF !important;
        border: 2px dashed #C8E6C9 !important;
        border-radius: 12px !important;
        min-height: 180px !important;
        padding: 2rem !important;
    }

    [data-testid="stFileUploaderDropzone"]:hover,
    [data-testid="stFileUploadDropzone"]:hover {
        background-color: #F8F9FA !important;
        border-color: #7BA591 !important;
    }

    [data-testid="stFileUploaderDropzoneInstructions"],
    [data-testid="stFileUploadDropzone"] span,
    [data-testid="stFileUploadDropzone"] p,
    [data-testid="stFileUploadDropzone"] small {
        color: #2C3E50 !important;
        font-size: 1.1rem !important;
        font-weight: 600 !important;
    }

    /* 업로드된 파일 목록 */
    [data-testid="stFileUploaderFile"],
    [data-testid="stFileUploaderFileName"] {
        background-color: white !important;
        color: #1E293B !important;
    }

    /* 파일 업로더 버튼 */
    [data-testid="stFileUploaderButton"] button,
    [data-testid="stFileUploadDropzone"] button {
        background-color: #81C784 !important;
        color: white !important;
        border: none !important;
    }

    /* 채팅 메시지 */
    .stChatMessage {
        background-color: white;
        border-radius: 12px;
        padding: 1rem;
        margin: 0.5rem 0;
        border: 1px solid #E5E7EB;
    }

    .stChatMessage[data-testid="user-message"] {
        background-color: #DBEAFE;
        border-color: #C8E6C9;
    }

    .stChatMessage[data-testid="assistant-message"] {
        background-color: #F8F9FA;
        border-color: #E5E7EB;
    }
</style>

<script>
// 페이지 로드 시 모든 데이터프레임을 강제로 라이트 모드로 변경
function forceLightTheme() {
    // 모든 데이터프레임 찾기
    const dataframes = document.querySelectorAll('[data-testid="stDataFrame"]');
    dataframes.forEach(df => {
        df.style.backgroundColor = 'white';
        df.style.color = '#1E293B';

        // 내부 모든 요소도 라이트 모드로
        const allElements = df.querySelectorAll('*');
        allElements.forEach(el => {
            el.style.backgroundColor = 'white';
            el.style.color = '#1E293B';
        });

        // 테이블 행들
        const rows = df.querySelectorAll('tr');
        rows.forEach(row => {
            row.style.backgroundColor = 'white';
        });

        // 테이블 셀들
        const cells = df.querySelectorAll('td, th');
        cells.forEach(cell => {
            cell.style.backgroundColor = 'white';
            cell.style.color = '#1E293B';
        });
    });

    // 업로드 버튼도 수정
    const buttons = document.querySelectorAll('button');
    buttons.forEach(btn => {
        if (btn.style.backgroundColor === 'rgb(38, 39, 48)' ||
            btn.style.backgroundColor === '#262730') {
            btn.style.backgroundColor = '#81C784';
            btn.style.color = 'white';
        }
    });
}

// DOM 변경 감지하여 계속 적용
const observer = new MutationObserver(forceLightTheme);
observer.observe(document.body, { childList: true, subtree: true });

// 초기 실행
setTimeout(forceLightTheme, 100);
setTimeout(forceLightTheme, 500);
setTimeout(forceLightTheme, 1000);
</script>
""", unsafe_allow_html=True)


# 발주 이력 및 선택 초기화
if 'order_history' not in st.session_state:
    st.session_state.order_history = {}

if 'selected_items' not in st.session_state:
    st.session_state.selected_items = set()

if 'custom_quantities' not in st.session_state:
    st.session_state.custom_quantities = {}

# 데이터 로딩 함수
@st.cache_data  # 영구 캐시 (파일 경로가 같으면 계속 사용 - 매우 빠름)
def load_psi_data(file_path):
    """PSI 엑셀 파일 로딩"""
    if not os.path.exists(file_path):
        st.error(f"❌ PSI 파일을 찾을 수 없습니다: {file_path}")
        return None, None, None, None, None

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        st.error(f"❌ Excel 파일 읽기 실패: {e}")
        return None, None, None, None, None

    # 시트 목록 확인 (디버깅용)
    available_sheets = wb.sheetnames
    debug_info = [f"📋 발견된 시트: {', '.join(available_sheets)}"]

    # 대시보드 시트에서 직접 값 읽기
    dashboard_data = {
        'total_sku': 0,
        'total_value': 0,
        'avg_turnover_days': 30,
        'shortage': 0,
        'reorder': 0,
        'sheet_names': available_sheets,  # 디버깅용
    }

    # 대시보드 시트가 있으면 값 읽기
    if '대시보드' in wb.sheetnames:
        ws_dashboard = wb['대시보드']
        try:
            total_sku_val = ws_dashboard.cell(6, 3).value
            total_value_val = ws_dashboard.cell(7, 3).value
            avg_days_val = ws_dashboard.cell(8, 3).value
            shortage_val = ws_dashboard.cell(9, 3).value
            reorder_val = ws_dashboard.cell(10, 3).value

            debug_info.append(f"C6(SKU): {total_sku_val}")
            debug_info.append(f"C7(금액): {total_value_val}")
            debug_info.append(f"C8(일수): {avg_days_val}")

            if total_sku_val:
                dashboard_data['total_sku'] = int(total_sku_val)
            if total_value_val:
                dashboard_data['total_value'] = float(total_value_val)
            if avg_days_val:
                dashboard_data['avg_turnover_days'] = float(avg_days_val)
            if shortage_val:
                dashboard_data['shortage'] = int(shortage_val)
            if reorder_val:
                dashboard_data['reorder'] = int(reorder_val)
        except Exception as e:
            debug_info.append(f"오류: {e}")
            print(f"대시보드 값 읽기 오류: {e}")
            pass

    # 재고분석 데이터는 수식이므로, PSI_메인과 안전재고에서 직접 생성
    # 먼저 안전재고 시트에서 SKU 리스트 가져오기
    ws_safety_temp = wb['안전재고']
    sku_list = []
    for row in range(2, min(ws_safety_temp.max_row + 1, 410)):
        sku = ws_safety_temp.cell(row, 1).value
        if sku:
            sku_list.append(sku)

    # PSI_메인 시트에서 현재고 가져오기
    ws_psi_temp = wb['PSI_메인']
    psi_stock = {}
    for row in range(4, min(ws_psi_temp.max_row + 1, 412)):
        sku = ws_psi_temp.cell(row, 1).value
        if sku:
            # 열 7은 수식이므로 Linux에서 None 반환
            # 열 8 (H)에서 실제 데이터 읽기 (202512월 재고)
            stock = ws_psi_temp.cell(row, 8).value or 0
            psi_stock[sku] = stock

    # ABC-XYZ분석 시트에서 추가 정보 가져오기
    ws_abc_temp = wb['ABC-XYZ분석 (2)']
    abc_info = {}
    for row in range(2, min(ws_abc_temp.max_row + 1, 410)):
        sku = ws_abc_temp.cell(row, 3).value
        if sku:
            abc_info[sku] = {
                '제품명': ws_abc_temp.cell(row, 7).value,
                '카테고리': ws_abc_temp.cell(row, 5).value,  # 추정
                'ABC등급': ws_abc_temp.cell(row, 29).value,
            }

    # 재고분석 데이터 생성
    inventory_data = []
    for sku in sku_list:
        if sku in abc_info:
            inventory_data.append({
                '구분': '정상',
                'SKU코드': str(sku),
                '제품명': abc_info[sku].get('제품명'),
                '카테고리': abc_info[sku].get('카테고리', '미분류'),
                'ABC등급': abc_info[sku].get('ABC등급'),
                'XYZ등급': None,  # 나중에 안전재고에서 병합
                '현재고': psi_stock.get(sku, 0),
                '안전재고': 0,  # 나중에 안전재고에서 업데이트
                '최근3개월평균': 0,
            })

    df_inventory = pd.DataFrame(inventory_data)

    # 안전재고 데이터
    ws_safety = wb['안전재고']
    safety_data = []
    for row in range(2, min(ws_safety.max_row + 1, 410)):
        sku = ws_safety.cell(row, 1).value
        if sku:
            # MOQ와 공급업체는 현재 PSI 파일에 없으므로 기본값 사용
            # 향후 별도 시트 추가 시 로드 가능
            safety_data.append({
                'SKU코드': sku,
                '제품명': ws_safety.cell(row, 1).value,
                '리드타임': ws_safety.cell(row, 3).value or 30,
                '일평균판매': ws_safety.cell(row, 4).value or 0,
                '수요표준편차': ws_safety.cell(row, 5).value or 0,
                'ABC': ws_safety.cell(row, 6).value,
                'XYZ': ws_safety.cell(row, 7).value,
                '안전재고': ws_safety.cell(row, 9).value or 0,
                'MOQ': None,  # 기본값: MOQ 없음 (필요시 PSI 파일에 컬럼 추가)
                '공급업체': '미지정',  # 기본값 (필요시 PSI 파일에 컬럼 추가)
            })

    df_safety = pd.DataFrame(safety_data)

    # ABC-XYZ 데이터 (월별 출고 포함)
    ws_abc = wb['ABC-XYZ분석 (2)']
    abc_data = []
    for row in range(2, min(ws_abc.max_row + 1, 410)):
        sku = ws_abc.cell(row, 3).value  # 컬럼 3: SKU#
        if sku:
            row_data = {
                'SKU코드': sku,
                '제품명': ws_abc.cell(row, 7).value,     # 컬럼 7: 제품명
                '연간판매': ws_abc.cell(row, 22).value or 0,  # 컬럼 22: 연간 판매
                '매입원가': ws_abc.cell(row, 24).value or 0,  # 컬럼 24: 평균 판매단가
                '연간COGS': ws_abc.cell(row, 26).value or 0,  # 컬럼 26: 25년 연간 판매금액
                'ABC등급': ws_abc.cell(row, 29).value,   # 컬럼 29: ABC등급
            }

            # 월별 출고 데이터 (컬럼 10-21: 1월~12월)
            for month_idx in range(1, 13):
                col_idx = 9 + month_idx  # 10-21 컬럼
                month_value = ws_abc.cell(row, col_idx).value or 0
                row_data[f'{month_idx}월출고'] = month_value

            abc_data.append(row_data)

    df_abc = pd.DataFrame(abc_data)

    # PSI 메인 데이터
    ws_psi = wb['PSI_메인']
    psi_data = []
    for row in range(4, min(ws_psi.max_row + 1, 412)):
        sku = ws_psi.cell(row, 1).value
        if sku:
            psi_data.append({
                'SKU코드': sku,
                '제품명': ws_psi.cell(row, 1).value,
                '카테고리': ws_psi.cell(row, 3).value,
                '계절': ws_psi.cell(row, 4).value,
                'ABC등급': ws_psi.cell(row, 5).value,
                'XYZ등급': ws_psi.cell(row, 6).value,
                '기초재고': ws_psi.cell(row, 7).value or 0,
            })

    df_psi = pd.DataFrame(psi_data)

    # SKU코드를 모두 문자열로 통일 (데이터 타입 충돌 방지)
    if 'SKU코드' in df_inventory.columns:
        df_inventory['SKU코드'] = df_inventory['SKU코드'].astype(str)
    if 'SKU코드' in df_safety.columns:
        df_safety['SKU코드'] = df_safety['SKU코드'].astype(str)
    if 'SKU코드' in df_abc.columns:
        df_abc['SKU코드'] = df_abc['SKU코드'].astype(str)
    if 'SKU코드' in df_psi.columns:
        df_psi['SKU코드'] = df_psi['SKU코드'].astype(str)

    # 모든 숫자 컬럼을 숫자 타입으로 변환 (데이터 타입 에러 방지)
    numeric_cols_inventory = ['현재고', '안전재고', '발주점']
    for col in numeric_cols_inventory:
        if col in df_inventory.columns:
            df_inventory[col] = pd.to_numeric(df_inventory[col], errors='coerce').fillna(0)

    numeric_cols_safety = ['일평균판매', '리드타임', '안전재고']
    for col in numeric_cols_safety:
        if col in df_safety.columns:
            df_safety[col] = pd.to_numeric(df_safety[col], errors='coerce').fillna(0)

    numeric_cols_abc = ['연간COGS', '연간판매', '비중%']
    for col in numeric_cols_abc:
        if col in df_abc.columns:
            df_abc[col] = pd.to_numeric(df_abc[col], errors='coerce').fillna(0)

    numeric_cols_psi = ['판매', '입고', '기초재고']
    for col in numeric_cols_psi:
        if col in df_psi.columns:
            df_psi[col] = pd.to_numeric(df_psi[col], errors='coerce').fillna(0)

    # 대시보드 데이터를 실제 데이터로부터 항상 계산
    # Streamlit Cloud에서는 Excel 수식이 계산되지 않으므로 직접 계산 필수
    total_value = 0
    calc_method = "없음"
    debug_samples = {}

    try:
        # 방법 1: 재고분석 시트에서 직접 읽기 (현재고 + 매입원가 모두 있음!)
        debug_samples['sheet_check'] = {'재고분석_exists': '재고분석' in wb.sheetnames}

        if '재고분석' in wb.sheetnames:
            ws_inv_analysis = wb['재고분석']
            debug_samples['재고분석_info'] = {
                'max_row': ws_inv_analysis.max_row,
                'max_col': ws_inv_analysis.max_column,
                'header_col8': ws_inv_analysis.cell(1, 8).value,
                'header_col17': ws_inv_analysis.cell(1, 17).value
            }

            # 첫 3개 행의 실제 값 확인
            debug_samples['first_3_rows_raw'] = []
            for test_row in range(2, 5):
                debug_samples['first_3_rows_raw'].append({
                    'row': test_row,
                    'sku_col3': ws_inv_analysis.cell(test_row, 3).value,
                    'stock_col8': ws_inv_analysis.cell(test_row, 8).value,
                    'price_col17': ws_inv_analysis.cell(test_row, 17).value
                })

            inventory_calc_data = []

            for row in range(2, min(ws_inv_analysis.max_row + 1, 410)):
                sku = ws_inv_analysis.cell(row, 3).value  # Column 3: SKU코드
                if sku:
                    current_stock = ws_inv_analysis.cell(row, 8).value or 0  # Column 8: 현재고
                    unit_price = ws_inv_analysis.cell(row, 17).value or 0    # Column 17: 매입원가

                    inventory_calc_data.append({
                        'SKU코드': sku,
                        '현재고': current_stock,
                        '매입원가': unit_price,
                        '재고금액': float(current_stock) * float(unit_price) if current_stock and unit_price else 0
                    })

            debug_samples['calc_data_count'] = len(inventory_calc_data)

            if inventory_calc_data:
                df_inv_calc = pd.DataFrame(inventory_calc_data)
                df_inv_calc['현재고'] = pd.to_numeric(df_inv_calc['현재고'], errors='coerce').fillna(0)
                df_inv_calc['매입원가'] = pd.to_numeric(df_inv_calc['매입원가'], errors='coerce').fillna(0)
                df_inv_calc['재고금액'] = df_inv_calc['현재고'] * df_inv_calc['매입원가']

                total_value = df_inv_calc['재고금액'].sum()
                calc_method = "재고분석시트(열8×열17)"

                # 디버그 정보
                debug_samples['inventory_sample'] = df_inv_calc.head(3)[['SKU코드', '현재고']].to_dict('records')
                debug_samples['abc_sample'] = df_inv_calc.head(3)[['SKU코드', '매입원가']].to_dict('records')
                debug_samples['merged_after'] = df_inv_calc.head(3).to_dict('records')
                debug_samples['total_stock_sum'] = float(df_inv_calc['현재고'].sum())
                debug_samples['nonzero_stock_count'] = int((df_inv_calc['현재고'] > 0).sum())
                debug_samples['nonzero_price_count'] = int((df_inv_calc['매입원가'] > 0).sum())

                # 상위 5개
                df_temp_sorted = df_inv_calc.nlargest(5, '재고금액')
                dashboard_data['debug_top5'] = df_temp_sorted[['SKU코드', '현재고', '매입원가', '재고금액']].to_dict('records')

        # 방법 2: 원래 방법 (재고분석 시트가 없거나 실패한 경우)
        if total_value == 0 and len(df_inventory) > 0 and len(df_abc) > 0:
            # 디버그: 원본 데이터 샘플 확인
            debug_samples['inventory_columns'] = list(df_inventory.columns)
            debug_samples['abc_columns'] = list(df_abc.columns)
            debug_samples['inventory_sample'] = df_inventory.head(3)[['SKU코드', '현재고']].to_dict('records') if '현재고' in df_inventory.columns else []
            debug_samples['abc_sample'] = df_abc.head(3)[['SKU코드', '매입원가']].to_dict('records') if '매입원가' in df_abc.columns else []

            # 재고금액 = 현재고 × 매입원가 (올바른 계산 방법)
            if '매입원가' in df_abc.columns and '현재고' in df_inventory.columns:
                df_temp = pd.merge(
                    df_inventory[['SKU코드', '현재고']],
                    df_abc[['SKU코드', '매입원가']],
                    on='SKU코드',
                    how='left'
                )

                # 디버그: 병합 후 샘플
                debug_samples['merged_before'] = df_temp.head(3).to_dict('records')

                df_temp['현재고'] = pd.to_numeric(df_temp['현재고'], errors='coerce').fillna(0)
                df_temp['매입원가'] = pd.to_numeric(df_temp['매입원가'], errors='coerce').fillna(0)
                df_temp['재고금액'] = df_temp['현재고'] * df_temp['매입원가']

                # 디버그: 계산 후 샘플
                debug_samples['merged_after'] = df_temp.head(3)[['SKU코드', '현재고', '매입원가', '재고금액']].to_dict('records')
                debug_samples['total_stock_sum'] = float(df_temp['현재고'].sum())
                debug_samples['nonzero_stock_count'] = int((df_temp['현재고'] > 0).sum())
                debug_samples['nonzero_price_count'] = int((df_temp['매입원가'] > 0).sum())

                total_value = df_temp['재고금액'].sum()
                calc_method = "현재고×매입원가"

                # 디버그: 상위 5개 SKU 정보 저장
                df_temp_sorted = df_temp.nlargest(5, '재고금액')
                dashboard_data['debug_top5'] = df_temp_sorted[['SKU코드', '현재고', '매입원가', '재고금액']].to_dict('records')
            else:
                calc_method = "매입원가 또는 현재고 컬럼 없음"
    except Exception as e:
        # 에러 발생시 기본값 0 사용
        total_value = 0
        calc_method = f"에러: {str(e)}"
        debug_samples['error'] = str(e)

    # 계산 방법 및 디버그 정보 저장
    dashboard_data['calc_method'] = calc_method
    dashboard_data['debug_samples'] = debug_samples

    # dashboard_data가 비어있거나 0이면 계산한 값으로 업데이트
    # Streamlit Cloud에서는 Excel 수식이 계산 안 되므로 항상 계산 값 사용
    if dashboard_data['total_sku'] == 0 or dashboard_data['total_sku'] is None:
        dashboard_data['total_sku'] = len(df_inventory)

    # 계산한 값으로 항상 덮어쓰기 (Excel 수식 값 무시)
    dashboard_data['total_value'] = total_value

    return dashboard_data, df_inventory, df_safety, df_abc, df_psi

# 발주 필요 분석 함수
def analyze_procurement_needs(df_inventory, df_safety):
    """발주 필요 SKU 분석"""
    # 데이터 병합 (XYZ, MOQ, 공급업체 포함)
    df = pd.merge(df_inventory, df_safety[['SKU코드', '일평균판매', '리드타임', 'XYZ', 'MOQ', '공급업체']], on='SKU코드', how='left')

    # XYZ 컬럼명을 XYZ등급으로 변경
    if 'XYZ' in df.columns:
        df['XYZ등급'] = df['XYZ']

    # 빈 값 처리 및 타입 변환
    df['일평균판매'] = pd.to_numeric(df['일평균판매'], errors='coerce').fillna(0)
    df['리드타임'] = pd.to_numeric(df['리드타임'], errors='coerce').fillna(30)
    df['안전재고'] = pd.to_numeric(df['안전재고'], errors='coerce').fillna(0)
    df['현재고'] = pd.to_numeric(df['현재고'], errors='coerce').fillna(0)

    # 발주점 계산 (ROP = 일평균판매 × 리드타임 + 안전재고)
    df['발주점'] = (df['일평균판매'] * df['리드타임']) + df['안전재고']

    # 발주 필요 여부
    df['발주필요'] = df['현재고'] <= df['발주점']

    # 재고 상태
    def get_status(row):
        # 현재고가 0인 경우
        if row['현재고'] == 0:
            # 일평균판매가 있으면 긴급 부족
            if row['일평균판매'] > 0:
                return '🔴 부족'
            # 판매가 없으면 휴면 품목
            else:
                return '⚪ 휴면'

        # 안전재고가 없지만 현재고가 있는 경우
        if row['안전재고'] == 0:
            # 일평균판매 기준으로 7일 이하면 부족
            if row['일평균판매'] > 0 and row['현재고'] / row['일평균판매'] <= 7:
                return '🔴 부족'
            # 판매는 없지만 재고는 있음
            return '⚪ 휴면'

        # 정상 로직: 안전재고 대비 비율
        ratio = row['현재고'] / row['안전재고']
        if ratio < 1.0:
            return '🔴 부족'
        elif ratio < 1.5:
            return '🟡 재주문 필요'
        elif ratio <= 2.0:
            return '🟢 적정'
        else:
            return '🔵 과잉'

    df['재고상태'] = df.apply(get_status, axis=1)

    # 권장 발주량 계산
    def calc_order_qty(row):
        try:
            # 발주 불필요하면 0
            if not row['발주필요']:
                return 0

            # 일평균판매가 없으면 0
            if row['일평균판매'] <= 0:
                return 0

            # 부족분 계산 (발주점 - 현재고)
            shortage = row['발주점'] - row['현재고']

            # 현재고가 발주점보다 많으면 발주 불필요
            if shortage <= 0:
                return 0

            # 기본: 부족분 + 1주 판매량 (안전 마진)
            weekly_sales = float(row['일평균판매']) * 7
            base_qty = shortage + weekly_sales

            # 안전재고 자체가 이미 ABC/XYZ 보정이 적용된 값이므로
            # 추가 보정 없이 기본 발주량만 계산
            # (안전재고 → 발주점 → shortage 계산 시 이미 반영됨)

            final_qty = max(0, int(base_qty))

            # MOQ 적용 (MOQ가 있으면 MOQ의 배수로 올림)
            moq = row.get('MOQ')
            if moq and moq > 0:
                import math
                # 발주량이 MOQ보다 작으면 MOQ로 설정
                if final_qty < moq:
                    final_qty = int(moq)
                else:
                    # MOQ의 배수로 올림
                    final_qty = int(math.ceil(final_qty / moq) * moq)

            return final_qty
        except Exception as e:
            # 디버깅용: 에러 무시하지 말고 0 반환
            return 0

    df['권장발주량'] = df.apply(calc_order_qty, axis=1)

    # ===== 추가 기능: 재고 충분도 분석 =====
    # 재고 소진 예상일 계산
    df['재고소진일'] = df.apply(
        lambda row: int(row['현재고'] / row['일평균판매']) if row['일평균판매'] > 0 else 999,
        axis=1
    )

    # 충분도 상태
    def get_coverage_status(days):
        if days <= 7:
            return '🔴 위험 (7일 이하)'
        elif days <= 14:
            return '🟡 주의 (14일 이하)'
        elif days <= 30:
            return '🟢 양호 (30일 이하)'
        elif days < 999:
            return '🔵 과다 (30일 초과)'
        else:
            return '⚪ 판매없음'

    df['충분도상태'] = df['재고소진일'].apply(get_coverage_status)

    # 리드타임 대비 안전도 (재고일 / 리드타임)
    df['리드타임대비'] = df.apply(
        lambda row: round(row['재고소진일'] / row['리드타임'], 1) if row['리드타임'] > 0 else 0,
        axis=1
    )

    # 발주 필요 여부 재계산 (재고 소진일 고려)
    # 재고가 30일 이상 있으면 발주 불필요로 변경
    df['발주필요'] = df.apply(
        lambda row: row['발주필요'] and row['재고소진일'] < 30,
        axis=1
    )

    return df

# 메인 앱
def main():
    # ==================== Deepflow 스타일 사이드바 ====================
    with st.sidebar:
        # 로고 및 브랜드
        st.markdown("""
        <div style='text-align: center; padding: 1.5rem 0 1rem 0;'>
            <div style='font-size: 1.5rem; font-weight: 700; color: #C8E6C9; margin-bottom: 0.25rem;'>
                📦 SmartFlow
            </div>
            <div style='font-size: 0.75rem; color: #64748B; letter-spacing: 1px;'>
                PROCUREMENT SYSTEM
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='height: 0.5rem;'></div>", unsafe_allow_html=True)

        # ===== SOLUTIONS 섹션 =====
        st.markdown("""
        <div style='color: #94A3B8; font-size: 0.75rem; font-weight: 600;
                    text-transform: uppercase; letter-spacing: 1px; margin: 1rem 0 0.5rem 0.5rem;'>
            Solutions
        </div>
        """, unsafe_allow_html=True)

        # 현재 선택된 탭을 session_state로 관리
        if 'current_page' not in st.session_state:
            st.session_state.current_page = "대시보드"

        # 파일 선택
        CURRENT_PSI_FILE_CHECK = "current_psi.xlsx"

        # 파일 업로드 직후 강제로 "마지막 업로드 파일" 모드로 전환
        if 'just_uploaded' not in st.session_state:
            st.session_state.just_uploaded = False

        with st.expander("☁️ 데이터 소스 선택", expanded=True):
            # 방금 업로드했거나 파일이 있으면 "마지막 업로드 파일"을 기본값으로
            if st.session_state.just_uploaded or os.path.exists(CURRENT_PSI_FILE_CHECK):
                default_index = 0
                st.session_state.just_uploaded = False  # 플래그 리셋
            else:
                default_index = 1

            file_option = st.radio(
                "파일 옵션:",
                ["마지막 업로드 파일", "파일 업로드"],
                index=default_index,
                label_visibility="collapsed",
                horizontal=False,
                key='file_option_radio'
            )

    # 헤더
    st.markdown(f"""
    <div style='padding: 1rem 0 0.5rem 0;'>
        <h1 style='color: #1E293B; font-size: 1.75rem; font-weight: 700; margin: 0;'>
            {st.session_state.current_page}
        </h1>
        <p style='color: #64748B; font-size: 0.875rem; margin: 0.25rem 0 0 0;'>
            분석 기준일: {datetime.now().strftime('%Y년 %m월 %d일')}
        </p>
    </div>
    """, unsafe_allow_html=True)

    excel_file = None
    CURRENT_PSI_FILE = "current_psi.xlsx"

    if file_option == "마지막 업로드 파일":
        if os.path.exists(CURRENT_PSI_FILE):
            excel_file = CURRENT_PSI_FILE
            # 파일 수정 시간 및 경로 확인
            abs_read_path = os.path.abspath(CURRENT_PSI_FILE)
            mtime = os.path.getmtime(CURRENT_PSI_FILE)
            mtime_str = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
            file_size = os.path.getsize(CURRENT_PSI_FILE)
            st.sidebar.success(f"✅ 마지막 업로드 파일 사용 중\n📅 {mtime_str}\n📁 {abs_read_path}\n📦 {file_size} bytes")
        else:
            st.sidebar.warning("⚠️ 업로드된 파일이 없습니다\n먼저 '파일 업로드'를 사용하세요")

    elif file_option == "기본 파일":
        excel_file = "PSI_260205_신규.xlsx"
        st.sidebar.success("✅ 기본 파일 사용 중")

    elif file_option == "파일 업로드":
        st.sidebar.markdown("""
        <div style='background: #C8E6C9; padding: 1rem; border-radius: 8px; margin-bottom: 1rem;'>
            <p style='color: #2C3E50; font-weight: 600; margin: 0; text-align: center;'>
                📤 PSI 파일 업로드
            </p>
        </div>
        """, unsafe_allow_html=True)

        uploaded_file = st.sidebar.file_uploader(
            "엑셀 파일 선택",
            type=['xlsx'],
            help="PSI_최종완성.xlsx 형식의 파일을 업로드하세요",
            label_visibility="collapsed"
        )
        if uploaded_file:
            # 업로드된 파일을 current_psi.xlsx로 저장
            abs_save_path = os.path.abspath(CURRENT_PSI_FILE)
            with open(CURRENT_PSI_FILE, 'wb') as f:
                f.write(uploaded_file.getvalue())

            # 저장 확인
            if os.path.exists(CURRENT_PSI_FILE):
                file_size = os.path.getsize(CURRENT_PSI_FILE)
                st.sidebar.info(f"📁 저장 위치: {abs_save_path}\n📦 크기: {file_size} bytes")

            # 업로드 완료 플래그 설정
            st.session_state.just_uploaded = True

            # 수식 캐시 생성 (Windows에서만 실행)
            import platform
            if platform.system() == 'Windows':
                with st.spinner('📊 수식 계산 중... (10초 소요)'):
                    try:
                        import win32com.client
                        excel = win32com.client.Dispatch("Excel.Application")
                        excel.Visible = False
                        excel.DisplayAlerts = False
                        abs_path = os.path.abspath(CURRENT_PSI_FILE)
                        workbook = excel.Workbooks.Open(abs_path)
                        excel.Calculation = -4105
                        excel.CalculateFull()
                        for sheet in workbook.Worksheets:
                            sheet.Calculate()
                        excel.CalculateFull()
                        workbook.Save()
                        workbook.Close(SaveChanges=True)
                        excel.Quit()
                        st.sidebar.success(f"✅ {uploaded_file.name} 업로드 완료\n📊 수식 캐시 생성 완료")
                    except Exception as e:
                        st.sidebar.success(f"✅ {uploaded_file.name} 업로드 완료")
                    finally:
                        st.rerun()
            else:
                # Linux/Streamlit Cloud - 수식 계산 건너뛰기
                st.sidebar.success(f"✅ {uploaded_file.name} 업로드 완료!")
                import time
                time.sleep(0.1)  # 파일 쓰기 완료 대기
                st.rerun()

            excel_file = CURRENT_PSI_FILE
        else:
            st.sidebar.info("파일을 업로드하세요")

    elif file_option == "경로 입력":
        custom_path = st.sidebar.text_input(
            "파일 경로:",
            value="PSI_260205_신규.xlsx",
            help="예: C:/data/PSI.xlsx"
        )
        if custom_path:
            if os.path.exists(custom_path):
                excel_file = custom_path
                st.sidebar.success(f"✅ 파일 찾음")
            else:
                st.sidebar.error("❌ 파일 없음")

    # 데이터 로딩
    if excel_file:
        with st.spinner('PSI 데이터 로딩 중...'):
            dashboard_data, df_inventory, df_safety, df_abc, df_psi = load_psi_data(excel_file)
            # PSI 파일 경로 저장 (발주 기록용)
            st.session_state.psi_file_path = excel_file

            # 디버그 정보 표시
            if dashboard_data:
                with st.sidebar.expander("📊 데이터 로딩 정보", expanded=True):
                    st.write(f"✅ SKU 수: {dashboard_data.get('total_sku', 0)}개")
                    st.write(f"✅ 재고금액: {dashboard_data.get('total_value', 0):,.0f}원")
                    st.write(f"✅ 평균일: {dashboard_data.get('avg_turnover_days', 0):.1f}일")
                    st.write(f"✅ 계산방법: {dashboard_data.get('calc_method', '없음')}")
                    if len(df_abc) > 0:
                        st.write(f"✅ ABC 데이터: {len(df_abc)}행")
                    if len(df_inventory) > 0:
                        st.write(f"✅ 재고 데이터: {len(df_inventory)}행")

                    # 시트 이름 표시
                    if 'sheet_names' in dashboard_data:
                        st.write(f"📋 **Excel 시트:** {', '.join(dashboard_data['sheet_names'][:3])}")

                    # 상세 디버그 정보
                    debug_samples = dashboard_data.get('debug_samples', {})
                    if debug_samples:
                        st.write("---")
                        st.write("**🔍 상세 디버깅:**")

                        if 'nonzero_stock_count' in debug_samples:
                            st.write(f"- 재고 있는 SKU: {debug_samples['nonzero_stock_count']}개")
                        if 'nonzero_price_count' in debug_samples:
                            st.write(f"- 가격 있는 SKU: {debug_samples['nonzero_price_count']}개")
                        if 'total_stock_sum' in debug_samples:
                            st.write(f"- 총 재고수량: {debug_samples['total_stock_sum']:,.0f}개")

                        if 'inventory_sample' in debug_samples and debug_samples['inventory_sample']:
                            st.write("**재고 샘플 (상위 3개):**")
                            for item in debug_samples['inventory_sample']:
                                st.write(f"  {item}")

                        if 'abc_sample' in debug_samples and debug_samples['abc_sample']:
                            st.write("**매입원가 샘플 (상위 3개):**")
                            for item in debug_samples['abc_sample']:
                                st.write(f"  {item}")

                        if 'merged_after' in debug_samples and debug_samples['merged_after']:
                            st.write("**병합 후 샘플 (상위 3개):**")
                            for item in debug_samples['merged_after']:
                                st.write(f"  {item}")

                    # 상위 5개 SKU 재고금액 표시
                    if 'debug_top5' in dashboard_data and dashboard_data['debug_top5']:
                        st.write("---")
                        st.write("**📦 재고금액 상위 5개 SKU:**")
                        for item in dashboard_data['debug_top5']:
                            st.write(f"- {item['SKU코드']}: {item['현재고']:.0f}개 × {item['매입원가']:,.0f}원 = {item['재고금액']:,.0f}원")
    else:
        dashboard_data, df_inventory, df_safety, df_abc, df_psi = None, None, None, None, None

    if dashboard_data is None:
        st.stop()

    # 발주 분석
    df_analysis = analyze_procurement_needs(df_inventory, df_safety)

    # df_abc와 merge하여 매입원가 추가
    if df_abc is not None and len(df_abc) > 0:
        df_analysis = df_analysis.merge(
            df_abc[['SKU코드', '매입원가']],
            on='SKU코드',
            how='left'
        )
        df_analysis['매입원가'] = df_analysis['매입원가'].fillna(0)

    # df_analysis를 session_state에 저장 (일일 리포트용)
    st.session_state.df_analysis = df_analysis

    # ===== 사이드바 계속 =====
    with st.sidebar:
        # 현재 파일 정보 표시
        if os.path.exists("current_psi.xlsx"):
            mtime = os.path.getmtime("current_psi.xlsx")
            mtime_str = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M')
            st.markdown(f"""
            <div style='font-size: 0.75rem; color: #64748B; padding: 0.5rem;
                        background-color: #F8F9FA; border-radius: 6px; margin-top: 0.5rem;'>
                📄 현재 파일<br/>
                <span style='color: #1E293B;'>📅 {mtime_str}</span>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<div style='height: 2rem;'></div>", unsafe_allow_html=True)

    # 사이드바 - 필터
    st.sidebar.markdown("""
    <div style='color: #94A3B8; font-size: 0.75rem; font-weight: 600;
                text-transform: uppercase; letter-spacing: 1px; margin: 0 0 0.5rem 0.5rem;'>
        Filters
    </div>
    """, unsafe_allow_html=True)

    # ABC 필터
    abc_filter = st.sidebar.multiselect(
        "ABC 등급",
        options=['A', 'B', 'C'],
        default=['A', 'B', 'C']
    )

    # 상태 필터
    status_filter = st.sidebar.multiselect(
        "재고 상태",
        options=['🔴 부족', '🟡 재주문 필요', '🟢 적정', '🔵 과잉'],
        default=['🔴 부족', '🟡 재주문 필요']
    )

    # 필터 적용
    df_filtered = df_analysis[
        (df_analysis['ABC등급'].isin(abc_filter)) &
        (df_analysis['재고상태'].isin(status_filter))
    ]

    # 탭 구성
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
        "📊 대시보드",
        "📦 발주 관리",
        "📋 발주 현황",
        "🎯 KPI 관리",
        "📈 분석",
        "🎲 시뮬레이션",
        "⚡ 자동 발주",
        "⚙️ 설정"
    ])

    with tab1:
        show_dashboard(dashboard_data, df_analysis)

    with tab2:
        show_procurement(df_filtered)

    with tab3:
        show_order_status(df_analysis)

    with tab4:
        show_kpi_management(df_analysis, df_psi, df_abc)

    with tab5:
        show_analysis(df_analysis, df_abc)

    with tab6:
        show_simulation(df_analysis, df_psi)

    with tab7:
        show_auto_orders(df_analysis)

    with tab8:
        show_settings()

def show_dashboard(dashboard_data, df_analysis):
    """대시보드 화면"""

    # 헤더 - 깔끔한 회색 톤
    st.markdown("""
        <div style='background: #ffffff; padding: 1.5rem; border-radius: 8px;
                    box-shadow: 0 1px 3px rgba(0,0,0,0.1); margin-bottom: 2rem; border: 1px solid #e2e8f0;'>
            <h2 style='color: #0f172a; margin: 0; font-size: 1.75rem; font-weight: 700;'>
                📊 재고 현황 대시보드
            </h2>
            <p style='color: #64748b; margin-top: 0.5rem; font-size: 0.95rem;'>
                실시간 재고 모니터링 및 스마트 발주 인사이트
            </p>
        </div>
    """, unsafe_allow_html=True)

    # 알림 배너 (긴급 상황 시)
    risk_count = len(df_analysis[df_analysis['충분도상태'].str.contains('위험', na=False)])
    order_needed_count = len(df_analysis[(df_analysis['발주필요'] == True) & (df_analysis['권장발주량'] > 0)])

    if risk_count > 0 or order_needed_count >= 10:
        if risk_count > 0:
            st.error(f"🚨 **긴급 알림**: {risk_count}개 품목이 재고 위험 상태입니다! (재고소진일 ≤7일)")
        if order_needed_count >= 10:
            st.warning(f"⚠️ **발주 알림**: {order_needed_count}개 품목이 발주가 필요합니다.")

    # 주요 지표 - 카드 스타일
    col1, col2, col3, col4, col5 = st.columns(5, gap="medium")

    with col1:
        total_sku = dashboard_data.get('total_sku', 0) or 0
        st.metric(
            label="총 SKU",
            value=f"{total_sku:,}개",
            delta=None
        )

    with col2:
        total_value = dashboard_data.get('total_value', 0) or 0
        st.metric(
            label="총 재고금액",
            value=f"{total_value/100000000:.1f}억원" if total_value > 0 else "0.0억원",
            delta=None
        )

    with col3:
        # 평균 재고 소진일 (신규 추가)
        avg_coverage = df_analysis[df_analysis['재고소진일'] < 999]['재고소진일'].mean()
        st.metric(
            label="📅 평균 재고일",
            value=f"{avg_coverage:.0f}일",
            delta="충분" if avg_coverage >= 14 else "부족",
            delta_color="normal" if avg_coverage >= 14 else "inverse"
        )

    with col4:
        # 재고 위험 제품 수 (신규 추가)
        risk_count = len(df_analysis[df_analysis['충분도상태'].str.contains('위험', na=False)])
        st.metric(
            label="🔴 재고 위험",
            value=f"{risk_count}개",
            delta="긴급 발주",
            delta_color="inverse" if risk_count > 0 else "normal"
        )

    with col5:
        avg_days = dashboard_data.get('avg_turnover_days', 30) or 30
        turnover_rate = 365 / avg_days if avg_days > 0 else 0
        st.metric(
            label="재고회전율",
            value=f"{turnover_rate:.1f}회/년",
            delta=f"목표 4회" if turnover_rate < 4 else "양호",
            delta_color="normal" if turnover_rate >= 4 else "inverse"
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # 긴급 조치 필요
    st.markdown("""
        <div style='background: #ffffff; padding: 1.25rem; border-radius: 8px;
                    box-shadow: 0 1px 3px rgba(0,0,0,0.1); margin-bottom: 1.5rem; border: 1px solid #e2e8f0;'>
            <h3 style='color: #dc2626; margin: 0; font-size: 1.25rem; font-weight: 700;'>
                🚨 긴급 조치 필요
            </h3>
        </div>
    """, unsafe_allow_html=True)

    urgent = df_analysis[df_analysis['재고상태'] == '🔴 부족'].sort_values('현재고', ascending=True).head(10)

    if len(urgent) > 0:
        for idx, row in urgent.iterrows():
            coverage_text = f"{row['재고소진일']}일치 재고" if row['재고소진일'] < 999 else "판매없음"

            # 순수 Streamlit 컴포넌트로 표시 (HTML 제거)
            sku_code = row['SKU코드']

            # 긴급 알림
            st.error(f"**⚠️ 즉시 발주 필요: {row['SKU코드']}**")
            st.write(f"**제품명**: {row['제품명']}")

            # 재고 정보
            urgent_col1, urgent_col2, urgent_col3 = st.columns(3)
            with urgent_col1:
                st.write(f"📦 현재고: **{row['현재고']:,.0f}개**")
            with urgent_col2:
                st.write(f"🛡️ 안전재고: **{row['안전재고']:,.0f}개**")
            with urgent_col3:
                st.write(f"📅 재고: **{coverage_text}**")

            # 최근 발주 정보
            last_order = st.session_state.order_history.get(sku_code, None)
            if last_order:
                days_ago = (datetime.now() - last_order['timestamp']).days
                st.caption(f"📋 최근발주: {last_order['quantity']:,.0f}개 ({days_ago}일 전)")
    else:
        st.markdown("""
            <div class="alert-success">
                <h4 style='margin: 0; font-size: 1.05rem; color: #ffffff; font-weight: 600;'>✅ 긴급 조치 필요한 품목이 없습니다</h4>
                <p style='margin: 0.375rem 0 0 0; color: #d1fae5; font-size: 0.9rem;'>모든 재고가 안전 수준을 유지하고 있습니다.</p>
            </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # 발주 대기
    st.markdown("""
        <div style='background: #ffffff; padding: 1.25rem; border-radius: 8px;
                    box-shadow: 0 1px 3px rgba(0,0,0,0.1); margin-bottom: 1.5rem; border: 1px solid #e2e8f0;'>
            <h3 style='color: #f59e0b; margin: 0; font-size: 1.25rem; font-weight: 700;'>
                📦 발주 필요 품목
            </h3>
        </div>
    """, unsafe_allow_html=True)

    # 실제로 발주가 필요한 품목만 (권장발주량 > 0)
    reorder = df_analysis[
        (df_analysis['발주필요'] == True) &
        (df_analysis['권장발주량'] > 0)
    ].sort_values('재고소진일', ascending=True).head(10)

    if len(reorder) > 0:
        # 전체 선택/해제
        col_select, col_clear, col_action = st.columns([1, 1, 3])
        with col_select:
            # 전체 선택 버튼
            if st.button("전체 선택", key="select_all_dashboard"):
                all_skus = set(reorder['SKU코드'].tolist())
                st.session_state.selected_items = all_skus
                # 체크박스 키 업데이트
                for idx, row in reorder.iterrows():
                    sku_code = row['SKU코드']
                    checkbox_key = f"check_{sku_code}_{idx}"
                    st.session_state[checkbox_key] = True

        with col_clear:
            if st.button("선택 해제", key="deselect_all"):
                st.session_state.selected_items = set()
                # 체크박스 키 업데이트
                for idx, row in reorder.iterrows():
                    sku_code = row['SKU코드']
                    checkbox_key = f"check_{sku_code}_{idx}"
                    st.session_state[checkbox_key] = False

        with col_action:
            selected_count = len(st.session_state.selected_items)
            if selected_count > 0:
                st.info(f"📦 선택된 품목: {selected_count}개")

        st.markdown("<br>", unsafe_allow_html=True)

        for idx, row in reorder.iterrows():
            col_check, col_content, col_qty, col_btn = st.columns([0.3, 2.7, 1, 1])

            with col_check:
                sku_code = row['SKU코드']
                is_checked = sku_code in st.session_state.selected_items
                checkbox_key = f"check_{sku_code}_{idx}"
                # 키가 없으면 is_checked로 초기화, 있으면 유지
                if checkbox_key not in st.session_state:
                    st.session_state[checkbox_key] = is_checked
                # 키 값과 is_checked 중 하나라도 True이면 True (전체 선택 반영)
                if is_checked or st.session_state.get(checkbox_key, False):
                    st.session_state[checkbox_key] = True

                checked = st.checkbox("", key=checkbox_key, label_visibility="collapsed")

                # 체크박스 변경 시 selected_items 업데이트
                if checked and sku_code not in st.session_state.selected_items:
                    st.session_state.selected_items.add(sku_code)
                elif not checked and sku_code in st.session_state.selected_items:
                    st.session_state.selected_items.remove(sku_code)

            with col_content:
                coverage_text = f"{row['재고소진일']}일치" if row['재고소진일'] < 999 else "충분"

                # 순수 Streamlit 컴포넌트로 표시 (HTML 제거)
                st.warning(f"**{row['SKU코드']} - {row['제품명']}**")

                # 발주 이력 확인
                sku_code = row['SKU코드']
                info_col1, info_col2, info_col3, info_col4 = st.columns(4)

                with info_col1:
                    st.write(f"📦 현재고: **{row['현재고']:,.0f}개**")
                with info_col2:
                    st.write(f"🛡️ 안전재고: **{row['안전재고']:,.0f}개**")
                with info_col3:
                    st.write(f"📅 {coverage_text}")
                with info_col4:
                    st.write(f"📋 권장: **{row['권장발주량']:,.0f}개**")

                # 최근 발주 정보
                last_order = st.session_state.order_history.get(sku_code, None)
                if last_order:
                    days_ago = (datetime.now() - last_order['timestamp']).days
                    st.caption(f"📋 최근발주: {last_order['quantity']:,.0f}개 ({days_ago}일 전)")

            with col_qty:
                # 발주 수량 입력
                default_qty = int(row['권장발주량'])
                if sku_code not in st.session_state.custom_quantities:
                    st.session_state.custom_quantities[sku_code] = default_qty

                qty = st.number_input(
                    "발주량",
                    min_value=1,
                    value=st.session_state.custom_quantities.get(sku_code, default_qty),
                    step=10,
                    key=f"qty_{sku_code}_{idx}",
                    label_visibility="collapsed"
                )
                st.session_state.custom_quantities[sku_code] = qty

            with col_btn:
                if st.button("📤 발주", key=f"order_{idx}", use_container_width=True):
                    # 발주 이력 저장 (사용자 입력 수량 사용)
                    order_qty = st.session_state.custom_quantities.get(sku_code, int(row['권장발주량']))
                    st.session_state.order_history[sku_code] = {
                        'quantity': order_qty,
                        'timestamp': datetime.now(),
                        'product_name': row['제품명']
                    }

                    # PSI 파일에 발주 기록
                    if 'psi_file_path' in st.session_state:
                        order_data = {
                            'SKU코드': sku_code,
                            '제품명': row['제품명'],
                            'ABC등급': row.get('ABC등급', 'N/A'),
                            'XYZ등급': row.get('XYZ등급', 'N/A'),
                            '현재고': row['현재고'],
                            '발주량': order_qty,
                            '매입원가': row.get('매입원가', 0),
                            '재고소진일': row.get('재고소진일', 0),
                            '일평균판매': row.get('일평균판매', 0),
                            '리드타임': row.get('리드타임', 30)
                        }
                        record_order_to_excel(st.session_state.psi_file_path, order_data)

                    st.success(f"✅ {sku_code} - {order_qty:,}개 발주 요청됨")
                    st.rerun()

        # 전체 선택/해제 플래그 초기화
        if 'just_selected_all_dash' in st.session_state:
            st.session_state.just_selected_all_dash = False
        if 'just_cleared_all_dash' in st.session_state:
            st.session_state.just_cleared_all_dash = False

        # 일괄 발주 버튼
        if len(st.session_state.selected_items) > 0:
            st.markdown("<br>", unsafe_allow_html=True)
            col1, col2, col3 = st.columns([1, 1, 2])

            with col1:
                if st.button(f"✅ 선택 품목 일괄 발주 ({len(st.session_state.selected_items)}개)", type="primary", use_container_width=True):
                    # 선택된 품목 발주 (사용자 입력 수량 사용)
                    total_qty = 0
                    for sku in st.session_state.selected_items:
                        matching_row = reorder[reorder['SKU코드'] == sku]
                        if len(matching_row) > 0:
                            row_data = matching_row.iloc[0]
                            # 사용자가 입력한 수량 사용, 없으면 권장 발주량
                            order_qty = st.session_state.custom_quantities.get(sku, int(row_data['권장발주량']))
                            st.session_state.order_history[sku] = {
                                'quantity': order_qty,
                                'timestamp': datetime.now(),
                                'product_name': row_data['제품명']
                            }
                            total_qty += order_qty

                            # PSI 파일에 발주 기록
                            if 'psi_file_path' in st.session_state:
                                order_data = {
                                    'SKU코드': sku,
                                    '제품명': row_data['제품명'],
                                    'ABC등급': row_data.get('ABC등급', 'N/A'),
                                    'XYZ등급': row_data.get('XYZ등급', 'N/A'),
                                    '현재고': row_data['현재고'],
                                    '발주량': order_qty,
                                    '매입원가': row_data.get('매입원가', 0),
                                    '재고소진일': row_data.get('재고소진일', 0),
                                    '일평균판매': row_data.get('일평균판매', 0),
                                    '리드타임': row_data.get('리드타임', 30)
                                }
                                record_order_to_excel(st.session_state.psi_file_path, order_data)

                    st.success(f"✅ {len(st.session_state.selected_items)}개 품목 발주 완료! (총 {total_qty:,}개)")
                    st.session_state.selected_items = set()
                    st.rerun()

            with col2:
                # Excel 발주서 다운로드
                from io import BytesIO
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "발주서"

                # 헤더
                ws['A1'] = '발주서'
                ws['A1'].font = Font(size=18, bold=True)
                ws['A2'] = f'발주일: {datetime.now().strftime("%Y-%m-%d %H:%M")}'

                # 컬럼 헤더
                headers = ['No', 'SKU코드', '제품명', '현재고', '안전재고', '발주량', '권장발주량']
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=4, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

                # 데이터
                row_num = 5
                for sku in st.session_state.selected_items:
                    matching_row = reorder[reorder['SKU코드'] == sku]
                    if len(matching_row) > 0:
                        row_data = matching_row.iloc[0]
                        order_qty = st.session_state.custom_quantities.get(sku, int(row_data['권장발주량']))
                        ws.cell(row=row_num, column=1, value=row_num-4)
                        ws.cell(row=row_num, column=2, value=sku)
                        ws.cell(row=row_num, column=3, value=row_data['제품명'])
                        ws.cell(row=row_num, column=4, value=int(row_data['현재고']))
                        ws.cell(row=row_num, column=5, value=int(row_data['안전재고']))
                        ws.cell(row=row_num, column=6, value=order_qty)
                        ws.cell(row=row_num, column=7, value=int(row_data['권장발주량']))
                        row_num += 1

                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                st.download_button(
                    label=f"📥 발주서 다운로드",
                    data=buffer,
                    file_name=f"발주서_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
    else:
        st.markdown("""
            <div class="alert-success">
                <h4 style='margin: 0; font-size: 1.05rem; color: #ffffff; font-weight: 600;'>✅ 재주문 필요한 품목이 없습니다</h4>
                <p style='margin: 0.375rem 0 0 0; color: #d1fae5; font-size: 0.9rem;'>현재 모든 재고가 적정 수준입니다.</p>
            </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # 재고 상태 분포
    st.markdown("""
        <div style='background: #ffffff; padding: 1.25rem; border-radius: 8px;
                    box-shadow: 0 1px 3px rgba(0,0,0,0.1); margin-bottom: 1.5rem; border: 1px solid #e2e8f0;'>
            <h3 style='color: #0f172a; margin: 0; font-size: 1.25rem; font-weight: 700;'>
                📊 재고 상태 분포
            </h3>
        </div>
    """, unsafe_allow_html=True)

    status_count = df_analysis['재고상태'].value_counts()

    # 전문적인 컬러 팔레트 (회색 계열)
    status_colors = ['#dc2626', '#f59e0b', '#059669', '#475569']

    fig = go.Figure(data=[go.Pie(
        labels=status_count.index,
        values=status_count.values,
        hole=.4,
        marker=dict(
            colors=status_colors,
            line=dict(color='white', width=3)
        ),
        textfont=dict(size=14, color='white', family='Arial Black'),
        pull=[0.1, 0, 0, 0]  # 첫 번째 항목 살짝 분리
    )])

    fig.update_layout(
        title=dict(
            text="재고 상태 분포",
            font=dict(size=16, color='#2C3E50', family='Arial')
        ),
        showlegend=True,
        height=400,
        paper_bgcolor='#FFFFFF',
        plot_bgcolor='#FFFFFF',
        font=dict(family='Arial, sans-serif', size=11, color='#5A6C7D')
    )

    st.plotly_chart(fig, use_container_width=True)

def show_procurement(df_filtered):
    """발주 관리 화면"""
    st.header("📦 스마트 발주 관리")

    # 발주 기준 설명 및 데이터 정보
    with st.expander("📋 발주 기준 및 계산 방식", expanded=False):
        col1, col2 = st.columns(2)

        with col1:
            st.markdown("""
            ### 🎯 발주 기준

            **발주 필요 판단**:
            - 현재고 ≤ 발주점
            - AND 재고소진일 < 30일

            **발주점 계산**:
            ```
            발주점 = (일평균판매 × 리드타임) + 안전재고
            ```
            ⚠️ **안전재고는 ABC/XYZ 등급에 따른 보정이 이미 적용된 값입니다**
            - ABC등급: 매출 기여도 (A > B > C)
            - XYZ등급: 수요 변동성 및 계절성 (Z = 높은 변동성)

            **권장 발주량**:
            ```
            발주량 = (발주점 - 현재고) + 1주 판매량
            ```
            부족분 + 안전 마진(1주치)

            **MOQ (최소발주량) 적용**:
            - MOQ가 설정된 경우, 발주량을 MOQ의 배수로 자동 조정
            - MOQ가 없는 경우, 권장 발주량 그대로 사용
            """)

        with col2:
            st.markdown("""
            ### 📊 데이터 정보

            **PSI 엑셀 파일에서 자동 계산**:
            - 안전재고 (ABC/XYZ 보정 적용됨)
            - 발주점 (리드타임 + 안전재고)
            - 일평균판매 (최근 30일 기준)
            - 재고소진일

            **ABC/XYZ 분석**:
            - ABC: 연간 매출 기여도 (파레토 법칙)
            - XYZ: 수요 변동계수 (표준편차/평균)
            - Z등급은 계절성 상품 포함

            **업데이트 방법**:
            1. 좌측 사이드바에서 파일 업로드
            2. 또는 경로 입력으로 자동 로드
            3. 일별/주별 업데이트 권장
            """)

        st.info("💡 **팁**: 발주량은 권장사항입니다. 실제 시장 상황, 프로모션 계획 등을 고려하여 조정하세요.")

        st.success("""
        ✅ **중요**: 안전재고는 PSI 엑셀 파일에서 이미 ABC/XYZ 등급별 보정계수가 적용되어 있습니다.
        - A등급/X등급: 높은 서비스 레벨 (안전재고 ↑)
        - Z등급: 계절성/변동성 반영 (안전재고 ↑)
        - C등급: 낮은 우선순위 (안전재고 ↓)
        """)

        st.warning("""
        ⚠️ **MOQ & 공급업체 기능**: 현재 PSI 파일에 MOQ(최소발주량)와 공급업체 정보가 없습니다.
        - 향후 별도 시트를 추가하시면 자동으로 반영됩니다
        - 현재는 모든 품목이 '공급업체: 미지정', 'MOQ: 없음'으로 표시됩니다
        """)

    st.markdown("---")

    # 발주 필요 품목 (권장발주량이 0보다 큰 것만)
    need_order = df_filtered[
        (df_filtered['발주필요'] == True) &
        (df_filtered['권장발주량'] > 0)
    ].sort_values('현재고', ascending=True)

    st.subheader(f"발주 필요 품목: {len(need_order)}개")

    if len(need_order) > 0:
        # Fragment 함수 정의 - 전체 선택/해제 버튼을 Fragment로 감싸서 스크롤 방지
        @st.fragment
        def render_select_buttons(need_order_df):
            # 전체 선택/해제
            col_select, col_clear = st.columns([1, 4])
            with col_select:
                # 전체 선택 버튼
                if st.button("전체 선택", key="select_all_reorder_tab"):
                    all_skus = set(need_order_df['SKU코드'].tolist())
                    st.session_state.selected_items = all_skus
                    # 체크박스 상태 업데이트 (각 체크박스 키에 맞게)
                    for enum_idx, (idx, row) in enumerate(need_order_df.iterrows()):
                        sku_code = row['SKU코드']
                        checkbox_key = f"sel_reorder_{sku_code}_{enum_idx}"
                        st.session_state[checkbox_key] = True

            with col_clear:
                if st.button("선택 해제", key="clear_all_reorder_tab"):
                    st.session_state.selected_items = set()
                    # 체크박스 상태 업데이트 (각 체크박스 키에 맞게)
                    for enum_idx, (idx, row) in enumerate(need_order_df.iterrows()):
                        sku_code = row['SKU코드']
                        checkbox_key = f"sel_reorder_{sku_code}_{enum_idx}"
                        st.session_state[checkbox_key] = False

            # 선택된 품목 수 표시
            selected_count = len(st.session_state.selected_items)
            if selected_count > 0:
                st.info(f"📦 선택된 품목: {selected_count}개")

        # 버튼 렌더링
        render_select_buttons(need_order)

        # Fragment 함수 정의 - 발주량 입력 부분만 rerun
        @st.fragment
        def render_order_item(row, enum_idx):
            sku_code = row['SKU코드']
            is_checked = sku_code in st.session_state.selected_items

            # 체크박스와 expander를 나란히 배치
            col_check, col_expand = st.columns([0.3, 4.7])

            with col_check:
                # 체크박스 키 관리
                checkbox_key = f"sel_reorder_{sku_code}_{enum_idx}"
                # 키가 없으면 is_checked로 초기화, 있으면 유지
                if checkbox_key not in st.session_state:
                    st.session_state[checkbox_key] = is_checked
                # 키 값과 is_checked 중 하나라도 True이면 True (전체 선택 반영)
                if is_checked or st.session_state.get(checkbox_key, False):
                    st.session_state[checkbox_key] = True

                # 체크박스
                selected = st.checkbox("선택", key=checkbox_key, label_visibility="collapsed")

                # 체크박스 변경 시 selected_items 업데이트
                if selected and sku_code not in st.session_state.selected_items:
                    st.session_state.selected_items.add(sku_code)
                elif not selected and sku_code in st.session_state.selected_items:
                    st.session_state.selected_items.remove(sku_code)

            with col_expand:
                # Expander 상태를 session_state로 관리
                expander_key = f"expander_{sku_code}_{enum_idx}"
                if expander_key not in st.session_state:
                    st.session_state[expander_key] = False

                with st.expander(
                    f"{'🔴' if row['재고상태'] == '🔴 부족' else '🟡'} {row['SKU코드']} - {row['제품명']}",
                    expanded=st.session_state[expander_key]
                ):
                    col1, col2, col3 = st.columns([2, 1, 1])

                    with col1:
                        st.write(f"**ABC/XYZ**: {row['ABC등급']}/{row['XYZ등급']}")
                        st.write(f"**현재고**: {row['현재고']:,.0f}개")
                        st.write(f"**안전재고**: {row['안전재고']:,.0f}개")
                        st.write(f"**발주점**: {row['발주점']:,.0f}개")
                        # MOQ 표시 (있는 경우에만)
                        if 'MOQ' in row and row['MOQ'] and row['MOQ'] > 0:
                            st.write(f"**MOQ**: {row['MOQ']:,.0f}개 (최소발주량)")

                    # 발주량 입력을 먼저 처리 (col3)
                    with col3:
                        # 발주 수량 입력 (session_state 사용)
                        if sku_code not in st.session_state.custom_quantities:
                            st.session_state.custom_quantities[sku_code] = int(row['권장발주량'])

                        order_qty = st.number_input(
                            "발주량",
                            min_value=0,
                            value=st.session_state.custom_quantities.get(sku_code, int(row['권장발주량'])),
                            step=100,
                            key=f"qty_reorder_{sku_code}_{enum_idx}"
                        )
                        st.session_state.custom_quantities[sku_code] = order_qty

                    # 발주량 입력 후 재고소진일 계산 (col2)
                    with col2:
                        st.write(f"**일평균 판매**: {row['일평균판매']:.1f}개")
                        st.write(f"**리드타임**: {row['리드타임']:.0f}일")

                        # 발주 전/후 재고 소진일 계산 (업데이트된 발주량 사용)
                        current_days = row['재고소진일']
                        order_qty = st.session_state.custom_quantities.get(sku_code, int(row['권장발주량']))
                        after_order_days = int((row['현재고'] + order_qty) / row['일평균판매']) if row['일평균판매'] > 0 else 0

                        st.write(f"**📅 발주 전 재고소진일**: {current_days}일")
                        st.write(f"**📅 발주 후 재고소진일**: {after_order_days}일")
                        st.write(f"**상태**: {row['충분도상태']}")
                        st.write(f"**예상 입고일**:")
                        expected_date = datetime.now() + timedelta(days=row['리드타임'])
                        st.write(expected_date.strftime('%Y-%m-%d'))

        # 품목별 발주
        selected_items = []

        for enum_idx, (idx, row) in enumerate(need_order.iterrows()):
            render_order_item(row, enum_idx)

            # session_state를 직접 확인하여 selected_items에 추가
            sku_code = row['SKU코드']
            if sku_code in st.session_state.selected_items:
                selected_items.append({
                    'SKU코드': row['SKU코드'],
                    '제품명': row['제품명'],
                    '발주량': st.session_state.custom_quantities.get(sku_code, int(row['권장발주량'])),
                    'ABC등급': row.get('ABC등급', 'N/A'),
                    'XYZ등급': row.get('XYZ등급', 'N/A'),
                    '현재고': row['현재고'],
                    '매입원가': row.get('매입원가', 0),
                    '재고소진일': row.get('재고소진일', 0),
                    '일평균판매': row.get('일평균판매', 0),
                    '리드타임': row.get('리드타임', 30),
                    'MOQ': row.get('MOQ'),
                    '공급업체': row.get('공급업체', '미지정')
                })

        # 전체 선택/해제 플래그 초기화
        if 'just_selected_all' in st.session_state:
            st.session_state.just_selected_all = False
        if 'just_cleared_all' in st.session_state:
            st.session_state.just_cleared_all = False

        st.markdown("---")

        # 발주 실행
        if len(selected_items) > 0:
            st.subheader(f"선택된 품목: {len(selected_items)}개")

            df_selected = pd.DataFrame(selected_items)
            st.dataframe(df_selected, use_container_width=True)

            col1, col2, col3 = st.columns([1, 1, 2])

            with col1:
                # Excel 발주서 다운로드
                from io import BytesIO
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill

                # 발주서 Excel 생성
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "발주서"

                # 헤더
                ws['A1'] = '발주서'
                ws['A1'].font = Font(size=18, bold=True)
                ws['A2'] = f'발주일: {datetime.now().strftime("%Y-%m-%d %H:%M")}'

                # 컬럼 헤더
                headers = ['No', 'SKU코드', '제품명', '공급업체', '발주량', 'MOQ', '비고']
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=4, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')

                # 데이터 (공급업체별로 정렬)
                sorted_items = sorted(selected_items, key=lambda x: x.get('공급업체', '미지정'))
                for row_idx, item in enumerate(sorted_items, start=5):
                    ws.cell(row=row_idx, column=1, value=row_idx-4)
                    ws.cell(row=row_idx, column=2, value=item['SKU코드'])
                    ws.cell(row=row_idx, column=3, value=item['제품명'])
                    ws.cell(row=row_idx, column=4, value=item.get('공급업체', '미지정'))
                    ws.cell(row=row_idx, column=5, value=item['발주량'])
                    ws.cell(row=row_idx, column=6, value=item.get('MOQ', ''))
                    ws.cell(row=row_idx, column=7, value='')

                # 컬럼 너비 조정
                ws.column_dimensions['A'].width = 5
                ws.column_dimensions['B'].width = 18
                ws.column_dimensions['C'].width = 35
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 10
                ws.column_dimensions['G'].width = 15

                # 바이트로 저장
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                st.download_button(
                    label="📥 발주서 Excel 다운로드",
                    data=buffer,
                    file_name=f"발주서_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )

            with col2:
                # 복사용 텍스트
                if st.button("📋 발주 내역 복사"):
                    order_text = f"📦 발주 요청\n\n발주일: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
                    order_text += "=" * 50 + "\n"
                    for idx, item in enumerate(selected_items, start=1):
                        order_text += f"{idx}. {item['SKU코드']} - {item['제품명']}\n"
                        order_text += f"   발주량: {item['발주량']}개\n\n"
                    order_text += "=" * 50 + "\n"
                    order_text += f"총 {len(selected_items)}개 품목"

                    st.text_area(
                        "아래 내용을 복사하세요",
                        value=order_text,
                        height=300
                    )

            with col3:
                # 전체 발주 실행
                if st.button("✅ 전체 발주 실행", type="primary", use_container_width=True):
                    total_qty = 0

                    # 모든 선택된 품목 발주
                    for item in selected_items:
                        sku = item['SKU코드']
                        qty = item['발주량']
                        st.session_state.order_history[sku] = {
                            'quantity': qty,
                            'timestamp': datetime.now(),
                            'product_name': item['제품명']
                        }
                        total_qty += qty

                        # PSI 파일에 발주 기록
                        if 'psi_file_path' in st.session_state:
                            order_data = {
                                'SKU코드': sku,
                                '제품명': item['제품명'],
                                'ABC등급': item.get('ABC등급', 'N/A'),
                                'XYZ등급': item.get('XYZ등급', 'N/A'),
                                '현재고': item.get('현재고', 0),
                                '발주량': qty,
                                '매입원가': item.get('매입원가', 0),
                                '재고소진일': item.get('재고소진일', 0),
                                '일평균판매': item.get('일평균판매', 0),
                                '리드타임': item.get('리드타임', 30)
                            }
                            record_order_to_excel(st.session_state.psi_file_path, order_data)

                    st.success(f"✅ 총 {len(selected_items)}개 품목, {total_qty:,}개 발주 완료!")
                    st.balloons()

                    # 선택 해제
                    st.session_state.selected_items = set()
                    st.rerun()
    else:
        st.success("✅ 현재 발주가 필요한 품목이 없습니다!")

def show_analysis(df_analysis, df_abc):
    """분석 화면"""
    st.header("📈 재고 분석")

    # ABC 분석
    st.subheader("ABC 등급별 분석")

    abc_summary = df_abc.groupby('ABC등급').agg({
        'SKU코드': 'count',
        '연간COGS': 'sum'
    }).reset_index()
    abc_summary.columns = ['ABC등급', 'SKU 수', '연간 COGS']
    abc_summary['비중%'] = abc_summary['연간 COGS'] / abc_summary['연간 COGS'].sum() * 100

    col1, col2 = st.columns(2)

    with col1:
        fig1 = px.bar(
            abc_summary,
            x='ABC등급',
            y='SKU 수',
            title='ABC 등급별 SKU 수',
            color='ABC등급',
            color_discrete_map={'A': '#dc2626', 'B': '#f59e0b', 'C': '#475569'}
        )
        fig1.update_layout(
            plot_bgcolor='#FFFFFF',
            paper_bgcolor='#FFFFFF',
            title_font=dict(size=15, color='#2C3E50', family='Arial'),
            xaxis=dict(showgrid=False, title_font=dict(color='#5A6C7D')),
            yaxis=dict(showgrid=True, gridcolor='#E5E7EB', title_font=dict(color='#5A6C7D')),
            font=dict(color='#5A6C7D')
        )
        fig1.update_traces(marker_line_width=0, textposition='outside')
        st.plotly_chart(fig1, use_container_width=True)

    with col2:
        fig2 = px.pie(
            abc_summary,
            values='연간 COGS',
            names='ABC등급',
            title='ABC 등급별 매출 비중',
            color='ABC등급',
            color_discrete_map={'A': '#dc2626', 'B': '#f59e0b', 'C': '#475569'},
            hole=0.3
        )
        fig2.update_layout(
            plot_bgcolor='#FFFFFF',
            paper_bgcolor='#FFFFFF',
            title_font=dict(size=15, color='#2C3E50', family='Arial'),
            font=dict(color='#5A6C7D')
        )
        fig2.update_traces(
            textfont=dict(size=12, color='white', family='Arial'),
            marker=dict(line=dict(color='#3a3a3a', width=2))
        )
        st.plotly_chart(fig2, use_container_width=True)

    # 테이블 표시용 포맷팅
    abc_display = abc_summary.copy()
    abc_display['연간 COGS'] = abc_display['연간 COGS'].apply(lambda x: f"{x:,.0f}")
    abc_display['비중%'] = abc_display['비중%'].apply(lambda x: f"{x:.1f}")

    st.dataframe(abc_display, use_container_width=True)

    st.markdown("---")

    # XYZ 분석
    st.subheader("XYZ 등급별 분석 (수요 변동성)")

    # XYZ 등급이 있는지 확인
    if 'XYZ등급' in df_analysis.columns:
        xyz_summary = df_analysis.groupby('XYZ등급').agg({
            'SKU코드': 'count',
            '현재고': 'sum'
        }).reset_index()
        xyz_summary.columns = ['XYZ등급', 'SKU 수', '총 재고량']
        xyz_summary['비중%'] = xyz_summary['SKU 수'] / xyz_summary['SKU 수'].sum() * 100

        col1, col2 = st.columns(2)

        with col1:
            fig_xyz1 = px.bar(
                xyz_summary,
                x='XYZ등급',
                y='SKU 수',
                title='XYZ 등급별 SKU 수',
                color='XYZ등급',
                color_discrete_map={'X': '#10b981', 'Y': '#f59e0b', 'Z': '#dc2626'}
            )
            fig_xyz1.update_layout(
                plot_bgcolor='#FFFFFF',
                paper_bgcolor='#FFFFFF',
                title_font=dict(size=15, color='#2C3E50', family='Arial'),
                xaxis=dict(showgrid=False, title_font=dict(color='#5A6C7D')),
                yaxis=dict(showgrid=True, gridcolor='#E5E7EB', title_font=dict(color='#5A6C7D')),
                font=dict(color='#5A6C7D')
            )
            fig_xyz1.update_traces(marker_line_width=0, textposition='outside')
            st.plotly_chart(fig_xyz1, use_container_width=True)

        with col2:
            fig_xyz2 = px.pie(
                xyz_summary,
                values='SKU 수',
                names='XYZ등급',
                title='XYZ 등급별 SKU 비중',
                color='XYZ등급',
                color_discrete_map={'X': '#10b981', 'Y': '#f59e0b', 'Z': '#dc2626'},
                hole=0.3
            )
            fig_xyz2.update_layout(
                plot_bgcolor='#FFFFFF',
                paper_bgcolor='#FFFFFF',
                title_font=dict(size=15, color='#2C3E50', family='Arial'),
                font=dict(color='#5A6C7D')
            )
            fig_xyz2.update_traces(
                textfont=dict(size=12, color='white', family='Arial'),
                marker=dict(line=dict(color='#3a3a3a', width=2))
            )
            st.plotly_chart(fig_xyz2, use_container_width=True)

        # 테이블 표시용 포맷팅
        xyz_display = xyz_summary.copy()
        xyz_display['총 재고량'] = xyz_display['총 재고량'].apply(lambda x: f"{x:,.0f}")
        xyz_display['비중%'] = xyz_display['비중%'].apply(lambda x: f"{x:.1f}")

        st.dataframe(xyz_display, use_container_width=True)

        # XYZ 등급 설명
        with st.expander("📖 XYZ 등급이란?"):
            st.markdown("""
            **XYZ 분석**은 수요의 변동성(예측 가능성)을 기준으로 재고를 분류합니다:

            - **X등급** 🟢: 변동성 낮음 (안정적 수요) - 예측이 쉬움
            - **Y등급** 🟡: 변동성 중간 - 예측이 보통
            - **Z등급** 🔴: 변동성 높음 (불안정 수요) - 예측이 어려움

            💡 **활용 팁**:
            - **AX** (높은 매출 + 안정적 수요): 최우선 재고 관리
            - **AZ** (높은 매출 + 불안정 수요): 안전재고 확보 필요
            - **CZ** (낮은 매출 + 불안정 수요): 최소 재고 유지
            """)

    st.markdown("---")

    # 재고회전 분석
    st.subheader("재고회전 분석")

    # 재고회전율 계산 (연간 판매 / 현재고)
    df_turnover = pd.merge(
        df_analysis[['SKU코드', '제품명', '현재고', 'ABC등급']],
        df_abc[['SKU코드', '연간판매']],
        on='SKU코드',
        how='left'
    )

    # 숫자 타입으로 변환
    df_turnover['연간판매'] = pd.to_numeric(df_turnover['연간판매'], errors='coerce').fillna(0)
    df_turnover['현재고'] = pd.to_numeric(df_turnover['현재고'], errors='coerce').fillna(0)

    df_turnover['재고회전율'] = df_turnover['연간판매'] / df_turnover['현재고'].replace(0, 1)
    df_turnover['재고회전일'] = 365 / df_turnover['재고회전율'].replace(0, 0.01)

    # 재고회전일을 숫자로 명시적 변환
    df_turnover['재고회전일'] = pd.to_numeric(df_turnover['재고회전일'], errors='coerce').fillna(999999)

    # TOP 10 느림 / 빠름
    col1, col2 = st.columns(2)

    with col1:
        st.write("**회전 느림 TOP 10** (개선 필요)")
        slow_turnover = df_turnover.nlargest(10, '재고회전일')[['SKU코드', '제품명', '재고회전일', 'ABC등급']]
        st.dataframe(slow_turnover, use_container_width=True)

    with col2:
        st.write("**회전 빠름 TOP 10**")
        fast_turnover = df_turnover[df_turnover['재고회전일'] > 0].nsmallest(10, '재고회전일')[['SKU코드', '제품명', '재고회전일', 'ABC등급']]
        st.dataframe(fast_turnover, use_container_width=True)

    st.markdown("---")

    # 재고회전일 상세 분석
    st.subheader("📊 재고회전일 상세 분석")

    # 재고회전일 구간 분류
    def classify_turnover_days(days):
        if days <= 30:
            return "0-30일 (우수)"
        elif days <= 60:
            return "31-60일 (양호)"
        elif days <= 90:
            return "61-90일 (보통)"
        else:
            return "90일 이상 (개선 필요)"

    df_turnover['회전구간'] = df_turnover['재고회전일'].apply(classify_turnover_days)

    # 구간별 통계
    col1, col2 = st.columns(2)

    with col1:
        # 구간별 SKU 수
        turnover_summary = df_turnover.groupby('회전구간').agg({
            'SKU코드': 'count'
        }).reset_index()
        turnover_summary.columns = ['회전구간', 'SKU 수']

        # 순서 정렬
        order = ["0-30일 (우수)", "31-60일 (양호)", "61-90일 (보통)", "90일 이상 (개선 필요)"]
        turnover_summary['회전구간'] = pd.Categorical(turnover_summary['회전구간'], categories=order, ordered=True)
        turnover_summary = turnover_summary.sort_values('회전구간')

        fig_turnover_bar = px.bar(
            turnover_summary,
            x='회전구간',
            y='SKU 수',
            title='재고회전일 구간별 SKU 수',
            color='회전구간',
            color_discrete_map={
                "0-30일 (우수)": '#10b981',
                "31-60일 (양호)": '#3b82f6',
                "61-90일 (보통)": '#f59e0b',
                "90일 이상 (개선 필요)": '#dc2626'
            }
        )
        fig_turnover_bar.update_layout(
            plot_bgcolor='#FFFFFF',
            paper_bgcolor='#FFFFFF',
            title_font=dict(size=15, color='#2C3E50', family='Arial'),
            xaxis=dict(showgrid=False, title_font=dict(color='#5A6C7D')),
            yaxis=dict(showgrid=True, gridcolor='#E5E7EB', title_font=dict(color='#5A6C7D')),
            font=dict(color='#5A6C7D'),
            showlegend=False
        )
        fig_turnover_bar.update_traces(marker_line_width=0, textposition='outside')
        st.plotly_chart(fig_turnover_bar, use_container_width=True)

    with col2:
        # 재고회전일 분포 히스토그램
        df_turnover_filtered = df_turnover[df_turnover['재고회전일'] < 365]  # 365일 이상 제외

        fig_turnover_hist = px.histogram(
            df_turnover_filtered,
            x='재고회전일',
            nbins=20,
            title='재고회전일 분포',
            color_discrete_sequence=['#3b82f6']
        )
        fig_turnover_hist.update_layout(
            plot_bgcolor='#FFFFFF',
            paper_bgcolor='#FFFFFF',
            title_font=dict(size=15, color='#2C3E50', family='Arial'),
            xaxis=dict(
                showgrid=False,
                title='재고회전일',
                title_font=dict(color='#5A6C7D')
            ),
            yaxis=dict(
                showgrid=True,
                gridcolor='#E5E7EB',
                title='SKU 수',
                title_font=dict(color='#5A6C7D')
            ),
            font=dict(color='#5A6C7D')
        )
        st.plotly_chart(fig_turnover_hist, use_container_width=True)

    # ABC 등급별 평균 재고회전일
    st.write("**ABC 등급별 평균 재고회전일**")
    abc_turnover = df_turnover.groupby('ABC등급').agg({
        '재고회전일': 'mean',
        'SKU코드': 'count'
    }).reset_index()
    abc_turnover.columns = ['ABC등급', '평균 재고회전일', 'SKU 수']
    abc_turnover['평균 재고회전일'] = abc_turnover['평균 재고회전일'].round(1)

    col1, col2, col3 = st.columns(3)
    for idx, row in abc_turnover.iterrows():
        with [col1, col2, col3][idx]:
            color = {'A': '🔴', 'B': '🟡', 'C': '⚫'}.get(row['ABC등급'], '⚪')
            st.metric(
                label=f"{color} {row['ABC등급']}등급 평균",
                value=f"{row['평균 재고회전일']:.1f}일",
                delta=f"{row['SKU 수']}개 SKU"
            )

    st.markdown("---")

    # 전체 상세 테이블
    with st.expander("📋 전체 SKU 재고회전일 상세 내역", expanded=False):
        st.write(f"총 {len(df_turnover)}개 SKU")

        # 정렬 옵션
        col1, col2 = st.columns([1, 3])
        with col1:
            sort_option = st.selectbox(
                "정렬 기준",
                ["재고회전일 느림순", "재고회전일 빠름순", "ABC등급", "SKU코드"]
            )

        # 정렬 적용
        if sort_option == "재고회전일 느림순":
            df_display = df_turnover.sort_values('재고회전일', ascending=False)
        elif sort_option == "재고회전일 빠름순":
            df_display = df_turnover[df_turnover['재고회전일'] > 0].sort_values('재고회전일', ascending=True)
        elif sort_option == "ABC등급":
            df_display = df_turnover.sort_values(['ABC등급', '재고회전일'], ascending=[True, False])
        else:
            df_display = df_turnover.sort_values('SKU코드')

        # 표시할 컬럼 선택
        display_columns = ['SKU코드', '제품명', 'ABC등급', '회전구간', '현재고', '연간판매', '재고회전율', '재고회전일']
        df_display = df_display[display_columns].copy()

        # 숫자 포맷팅
        df_display['현재고'] = df_display['현재고'].apply(lambda x: f"{x:,.0f}")
        df_display['연간판매'] = df_display['연간판매'].apply(lambda x: f"{x:,.0f}")
        df_display['재고회전율'] = df_display['재고회전율'].apply(lambda x: f"{x:.1f}")
        df_display['재고회전일'] = df_display['재고회전일'].apply(lambda x: f"{x:.1f}")

        # 테이블 표시
        st.dataframe(
            df_display,
            use_container_width=True,
            height=400
        )

        # 다운로드 버튼
        csv = df_display.to_csv(index=False, encoding='utf-8-sig')
        st.download_button(
            label="📥 CSV 다운로드",
            data=csv,
            file_name=f"재고회전일_상세_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv"
        )

def show_kpi_management(df_analysis, df_psi, df_abc):
    """KPI 관리 화면 - 월별 목표 설정 및 달성도 추적"""
    st.header("🎯 KPI 관리")

    # KPI 설명 추가
    with st.expander("📖 KPI 항목별 설명 및 기준", expanded=False):
        st.markdown("""
        ### 📊 5가지 핵심 KPI 설명

        #### 1️⃣ 재고회전율 (Inventory Turnover Rate)
        **정의**: 연간 재고가 몇 번 회전(판매)되는지를 나타내는 지표
        **계산**: (월간 출고량 / 평균 재고) × 12
        **업종별 기준**:
        - 제조업: 8~12회 (월 1회)
        - 유통업: 12~24회 (월 2회)
        - 식품/패션: 24~52회 (주 1회)

        **중요성**: 높을수록 재고 효율이 좋으며, 재고 보관 비용 절감
        ⚠️ 너무 높으면 품절 위험 증가

        ---

        #### 2️⃣ 평균재고소진일 (Average Days of Inventory)
        **정의**: 현재 재고가 완전히 소진되기까지 걸리는 평균 일수
        **계산**: 현재고 / 일평균판매
        **기준**:
        - 우수: 20~30일
        - 양호: 30~45일
        - 주의: 45~60일
        - 개선 필요: 60일 초과

        **중요성**: 낮을수록 재고 회전이 빠르고 신선도 유지

        ---

        #### 3️⃣ 발주정확도 (Order Accuracy)
        **정의**: 적정 시점에 적정 수량을 발주한 비율
        **계산**: (적정 발주 품목 수 / 전체 발주 품목 수) × 100
        **기준**:
        - 우수: 95% 이상
        - 양호: 90~95%
        - 개선 필요: 90% 미만

        **중요성**: 높을수록 재고 낭비 감소, 물류비 절감
        💡 재고소진일 10~30일 사이 발주를 적정으로 간주

        ---

        #### 4️⃣ 품절률 (Stockout Rate)
        **정의**: 전체 품목 중 재고 부족(🔴) 상태인 품목의 비율
        **계산**: (재고 부족 품목 수 / 전체 품목 수) × 100
        **기준**:
        - 우수: 2% 미만
        - 양호: 2~5%
        - 주의: 5~10%
        - 위험: 10% 초과

        **중요성**: 낮을수록 좋으며, 판매 기회 손실 방지
        💰 품절 1회 = 매출 손실 + 고객 신뢰 하락

        ---

        #### 5️⃣ 과다재고율 (Overstock Rate)
        **정의**: 전체 품목 중 과잉 재고(🔵) 상태인 품목의 비율
        **계산**: (과다 재고 품목 수 / 전체 품목 수) × 100
        **기준**:
        - 우수: 5% 미만
        - 양호: 5~10%
        - 주의: 10~15%
        - 개선 필요: 15% 초과

        **중요성**: 낮을수록 재고 자금 효율적 운영
        💸 과다재고 = 보관비 증가 + 자금 묶임 + 진부화 위험

        ---

        ### 🎯 균형 있는 KPI 관리가 핵심!
        - **품절률 ↓** vs **과다재고율 ↓**: 양쪽 균형이 중요
        - **재고회전율 ↑** vs **품절률 ↓**: 적정선 유지 필요
        - **발주정확도 ↑**: 모든 KPI 개선의 핵심
        """)

    # 현재 월
    current_month = datetime.now().strftime("%Y-%m")
    current_month_num = datetime.now().month  # 1-12

    # KPI 목표 설정 (실무에서는 엑셀이나 DB에서 가져오지만, 여기서는 session_state 사용)
    if 'kpi_targets' not in st.session_state:
        st.session_state.kpi_targets = {
            '재고회전율': 12.0,    # 연간 12회 (월 1회)
            '평균재고소진일': 30.0,  # 30일 이내
            '발주정확도': 95.0,      # 95%
            '품절률': 2.0,           # 2% 이하
            '과다재고율': 10.0        # 10% 이하
        }

    # 월별 목표 설정
    with st.expander("📝 월별 목표 설정", expanded=False):
        st.write(f"**현재 월**: {current_month}")

        col1, col2 = st.columns(2)

        with col1:
            target_turnover = st.number_input(
                "목표 재고회전율 (연간)",
                min_value=1.0,
                max_value=50.0,
                value=st.session_state.kpi_targets['재고회전율'],
                step=0.5,
                help="연간 재고회전율 목표 (12회 = 월 1회)"
            )
            st.session_state.kpi_targets['재고회전율'] = target_turnover

            target_days = st.number_input(
                "목표 평균재고소진일 (일)",
                min_value=1.0,
                max_value=180.0,
                value=st.session_state.kpi_targets['평균재고소진일'],
                step=1.0,
                help="재고가 소진되기까지 평균 일수"
            )
            st.session_state.kpi_targets['평균재고소진일'] = target_days

            target_accuracy = st.number_input(
                "목표 발주정확도 (%)",
                min_value=50.0,
                max_value=100.0,
                value=st.session_state.kpi_targets['발주정확도'],
                step=1.0,
                help="발주한 품목 중 정확하게 발주한 비율"
            )
            st.session_state.kpi_targets['발주정확도'] = target_accuracy

        with col2:
            target_stockout = st.number_input(
                "목표 품절률 (%)",
                min_value=0.0,
                max_value=50.0,
                value=st.session_state.kpi_targets['품절률'],
                step=0.5,
                help="재고 부족으로 판매 불가한 품목 비율"
            )
            st.session_state.kpi_targets['품절률'] = target_stockout

            target_overstock = st.number_input(
                "목표 과다재고율 (%)",
                min_value=0.0,
                max_value=50.0,
                value=st.session_state.kpi_targets['과다재고율'],
                step=1.0,
                help="재고가 과다한 품목 비율"
            )
            st.session_state.kpi_targets['과다재고율'] = target_overstock

    st.markdown("---")

    # 현재 실적 계산
    st.subheader(f"📊 {current_month} 실적 현황")

    # 1. 재고회전율 계산 (출고 데이터 기반)
    if len(df_abc) > 0 and len(df_analysis) > 0:
        # 당월 출고 데이터 합계
        month_col = f'{current_month_num}월출고'
        if month_col in df_abc.columns:
            total_monthly_sales = df_abc[month_col].sum()
        else:
            # 컬럼이 없으면 연간판매를 12로 나눈 평균 사용
            total_monthly_sales = df_abc['연간판매'].sum() / 12 if '연간판매' in df_abc.columns else 0

        # 평균 재고
        avg_inventory = df_analysis['현재고'].sum()

        if avg_inventory > 0 and total_monthly_sales > 0:
            # 월간 회전율 × 12 = 연간 회전율
            current_turnover = (total_monthly_sales / avg_inventory) * 12
        else:
            current_turnover = 0
    else:
        current_turnover = 0

    # 2. 평균 재고소진일
    avg_days = df_analysis['재고소진일'].replace([float('inf'), -float('inf')], 999).mean()
    if pd.isna(avg_days) or avg_days > 365:
        avg_days = 999

    # 3. 품절률 (재고 부족 품목 비율)
    total_skus = len(df_analysis)
    stockout_skus = len(df_analysis[df_analysis['재고상태'] == '🔴 부족'])
    stockout_rate = (stockout_skus / total_skus * 100) if total_skus > 0 else 0

    # 4. 과다재고율 (과다 재고 품목 비율)
    overstock_skus = len(df_analysis[df_analysis['재고상태'] == '🔵 과잉'])
    overstock_rate = (overstock_skus / total_skus * 100) if total_skus > 0 else 0

    # 5. 발주정확도 (발주 필요 품목 대비 적정 발주 비율 - 간략화)
    need_order_skus = len(df_analysis[df_analysis['발주필요'] == True])
    if need_order_skus > 0:
        # 재고소진일이 10-30일 사이인 것을 적정 발주로 간주
        proper_orders = len(df_analysis[
            (df_analysis['발주필요'] == True) &
            (df_analysis['재고소진일'] >= 10) &
            (df_analysis['재고소진일'] <= 30)
        ])
        order_accuracy = (proper_orders / need_order_skus * 100)
    else:
        order_accuracy = 100.0

    # KPI 카드 표시
    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        delta1 = current_turnover - st.session_state.kpi_targets['재고회전율']
        st.metric(
            label="재고회전율 (연간)",
            value=f"{current_turnover:.1f}회",
            delta=f"{delta1:+.1f}",
            delta_color="normal"
        )
        if current_turnover >= st.session_state.kpi_targets['재고회전율']:
            st.success("✅ 목표 달성")
        else:
            st.error(f"❌ 목표: {st.session_state.kpi_targets['재고회전율']:.1f}회")

    with col2:
        delta2 = -(avg_days - st.session_state.kpi_targets['평균재고소진일'])  # 낮을수록 좋음
        st.metric(
            label="평균 재고소진일",
            value=f"{avg_days:.1f}일",
            delta=f"{delta2:+.1f}일" if abs(delta2) < 100 else "N/A",
            delta_color="normal"
        )
        if avg_days <= st.session_state.kpi_targets['평균재고소진일']:
            st.success("✅ 목표 달성")
        else:
            st.error(f"❌ 목표: {st.session_state.kpi_targets['평균재고소진일']:.1f}일")

    with col3:
        delta3 = order_accuracy - st.session_state.kpi_targets['발주정확도']
        st.metric(
            label="발주 정확도",
            value=f"{order_accuracy:.1f}%",
            delta=f"{delta3:+.1f}%",
            delta_color="normal"
        )
        if order_accuracy >= st.session_state.kpi_targets['발주정확도']:
            st.success("✅ 목표 달성")
        else:
            st.error(f"❌ 목표: {st.session_state.kpi_targets['발주정확도']:.1f}%")

    with col4:
        delta4 = -(stockout_rate - st.session_state.kpi_targets['품절률'])  # 낮을수록 좋음
        st.metric(
            label="품절률",
            value=f"{stockout_rate:.1f}%",
            delta=f"{delta4:+.1f}%",
            delta_color="inverse"
        )
        if stockout_rate <= st.session_state.kpi_targets['품절률']:
            st.success("✅ 목표 달성")
        else:
            st.error(f"❌ 목표: {st.session_state.kpi_targets['품절률']:.1f}%")

    with col5:
        delta5 = -(overstock_rate - st.session_state.kpi_targets['과다재고율'])  # 낮을수록 좋음
        st.metric(
            label="과다재고율",
            value=f"{overstock_rate:.1f}%",
            delta=f"{delta5:+.1f}%",
            delta_color="inverse"
        )
        if overstock_rate <= st.session_state.kpi_targets['과다재고율']:
            st.success("✅ 목표 달성")
        else:
            st.error(f"❌ 목표: {st.session_state.kpi_targets['과다재고율']:.1f}%")

    st.markdown("---")

    # 월별 추이 (시뮬레이션 - 실제로는 DB에서 가져와야 함)
    st.subheader("📈 월별 KPI 추이")

    import plotly.graph_objects as go

    # 샘플 데이터 (실제로는 DB에서 가져와야 함)
    months = ['2024-09', '2024-10', '2024-11', '2024-12', '2025-01', current_month]
    turnover_history = [10.5, 11.2, 11.8, 12.3, 11.9, current_turnover]
    days_history = [35, 33, 31, 29, 31, avg_days]
    accuracy_history = [92, 93, 94, 95, 94, order_accuracy]

    tab_chart1, tab_chart2, tab_chart3 = st.tabs(["재고회전율", "재고소진일", "발주정확도"])

    with tab_chart1:
        fig1 = go.Figure()
        fig1.add_trace(go.Scatter(
            x=months,
            y=turnover_history,
            mode='lines+markers',
            name='실적',
            line=dict(color='#3b82f6', width=3)
        ))
        fig1.add_trace(go.Scatter(
            x=months,
            y=[st.session_state.kpi_targets['재고회전율']] * len(months),
            mode='lines',
            name='목표',
            line=dict(color='#2C3E50', width=2, dash='dash')
        ))
        fig1.update_layout(
            title='재고회전율 추이',
            xaxis_title='월',
            yaxis_title='회전율 (연간)',
            plot_bgcolor='#FFFFFF',
            paper_bgcolor='#FFFFFF',
            title_font=dict(color='#2C3E50'),
            xaxis=dict(color='#5A6C7D', gridcolor='#E5E7EB'),
            yaxis=dict(color='#5A6C7D', gridcolor='#E5E7EB'),
            font=dict(color='#5A6C7D'),
            height=400
        )
        st.plotly_chart(fig1, use_container_width=True)

    with tab_chart2:
        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(
            x=months,
            y=days_history,
            mode='lines+markers',
            name='실적',
            line=dict(color='#10b981', width=3)
        ))
        fig2.add_trace(go.Scatter(
            x=months,
            y=[st.session_state.kpi_targets['평균재고소진일']] * len(months),
            mode='lines',
            name='목표',
            line=dict(color='#2C3E50', width=2, dash='dash')
        ))
        fig2.update_layout(
            title='평균 재고소진일 추이',
            xaxis_title='월',
            yaxis_title='일수',
            plot_bgcolor='#FFFFFF',
            paper_bgcolor='#FFFFFF',
            title_font=dict(color='#2C3E50'),
            xaxis=dict(color='#5A6C7D', gridcolor='#E5E7EB'),
            yaxis=dict(color='#5A6C7D', gridcolor='#E5E7EB'),
            font=dict(color='#5A6C7D'),
            height=400
        )
        st.plotly_chart(fig2, use_container_width=True)

    with tab_chart3:
        fig3 = go.Figure()
        fig3.add_trace(go.Scatter(
            x=months,
            y=accuracy_history,
            mode='lines+markers',
            name='실적',
            line=dict(color='#f59e0b', width=3)
        ))
        fig3.add_trace(go.Scatter(
            x=months,
            y=[st.session_state.kpi_targets['발주정확도']] * len(months),
            mode='lines',
            name='목표',
            line=dict(color='#2C3E50', width=2, dash='dash')
        ))
        fig3.update_layout(
            title='발주 정확도 추이',
            xaxis_title='월',
            yaxis_title='정확도 (%)',
            plot_bgcolor='#FFFFFF',
            paper_bgcolor='#FFFFFF',
            title_font=dict(color='#2C3E50'),
            xaxis=dict(color='#5A6C7D', gridcolor='#E5E7EB'),
            yaxis=dict(color='#5A6C7D', gridcolor='#E5E7EB'),
            font=dict(color='#5A6C7D'),
            height=400
        )
        st.plotly_chart(fig3, use_container_width=True)

    st.markdown("---")

    # 개선 제안
    st.subheader("💡 개선 제안")

    improvement_suggestions = []

    if current_turnover < st.session_state.kpi_targets['재고회전율']:
        improvement_suggestions.append(
            f"⚠️ **재고회전율 개선 필요**: 현재 {current_turnover:.1f}회 → 목표 {st.session_state.kpi_targets['재고회전율']:.1f}회"
        )
        improvement_suggestions.append("   → 느린 회전 품목 재고 축소, 빠른 회전 품목 재고 확대")

    if avg_days > st.session_state.kpi_targets['평균재고소진일']:
        improvement_suggestions.append(
            f"⚠️ **재고소진일 단축 필요**: 현재 {avg_days:.1f}일 → 목표 {st.session_state.kpi_targets['평균재고소진일']:.1f}일"
        )
        improvement_suggestions.append("   → 과다 재고 품목 판촉, 발주량 조정")

    if order_accuracy < st.session_state.kpi_targets['발주정확도']:
        improvement_suggestions.append(
            f"⚠️ **발주정확도 향상 필요**: 현재 {order_accuracy:.1f}% → 목표 {st.session_state.kpi_targets['발주정확도']:.1f}%"
        )
        improvement_suggestions.append("   → 수요 예측 정확도 향상, 리드타임 관리 강화")

    if stockout_rate > st.session_state.kpi_targets['품절률']:
        improvement_suggestions.append(
            f"⚠️ **품절률 감소 필요**: 현재 {stockout_rate:.1f}% → 목표 {st.session_state.kpi_targets['품절률']:.1f}%"
        )
        improvement_suggestions.append("   → 안전재고 확보, 발주 타이밍 앞당기기")

    if overstock_rate > st.session_state.kpi_targets['과다재고율']:
        improvement_suggestions.append(
            f"⚠️ **과다재고율 감소 필요**: 현재 {overstock_rate:.1f}% → 목표 {st.session_state.kpi_targets['과다재고율']:.1f}%"
        )
        improvement_suggestions.append("   → 과다 재고 품목 판매 촉진, 발주량 축소")

    if len(improvement_suggestions) == 0:
        st.success("🎉 **모든 KPI 목표 달성!** 현재 재고 관리가 우수합니다.")
    else:
        for suggestion in improvement_suggestions:
            st.warning(suggestion)

def show_order_status(df_analysis):
    """발주 현황 대시보드 - PSI 파일에서 실제 발주 내역 읽기"""

    st.header("📋 발주 현황 대시보드")

    # PSI 파일에서 발주리스트 읽기
    psi_file = st.session_state.get('psi_file_path', 'current_psi.xlsx')
    df_orders = None

    if os.path.exists(psi_file):
        try:
            wb = openpyxl.load_workbook(psi_file, data_only=True)
            if '발주리스트' in wb.sheetnames:
                ws = wb['발주리스트']

                # 데이터 읽기 (헤더 제외)
                data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # 발주일이 있으면
                        data.append({
                            '발주일': row[0],
                            'SKU코드': row[1],
                            '제품명': row[2],
                            'ABC/XYZ': row[3],
                            '현재고': row[4],
                            '발주량': row[5],
                            '구매원가': row[6],
                            '발주 전 재고소진일': row[7],
                            '발주 후 재고소진일': row[8],
                            '예상입고일': row[9]
                        })

                if data:
                    df_orders = pd.DataFrame(data)
                wb.close()
        except Exception as e:
            st.error(f"발주 내역을 불러오는 중 오류: {str(e)}")

    # 발주 통계
    col1, col2, col3, col4 = st.columns(4)

    if df_orders is not None and len(df_orders) > 0:
        today = datetime.now().date()

        # 오늘 발주
        df_orders['발주일_date'] = pd.to_datetime(df_orders['발주일']).dt.date
        today_orders = df_orders[df_orders['발주일_date'] == today]

        with col1:
            st.metric("오늘 발주", f"{len(today_orders)}건")

        # 이번주 발주
        week_start = today - timedelta(days=today.weekday())
        week_orders = df_orders[df_orders['발주일_date'] >= week_start]

        with col2:
            st.metric("이번주 발주", f"{len(week_orders)}건")

        # 이번달 발주
        month_start = today.replace(day=1)
        month_orders = df_orders[df_orders['발주일_date'] >= month_start]

        with col3:
            st.metric("이번달 발주", f"{len(month_orders)}건")

        # 전체 발주
        with col4:
            st.metric("전체 발주", f"{len(df_orders)}건")
    else:
        with col1:
            st.metric("오늘 발주", "0건")
        with col2:
            st.metric("이번주 발주", "0건")
        with col3:
            st.metric("이번달 발주", "0건")
        with col4:
            st.metric("전체 발주", "0건")

    st.markdown("---")

    # 발주 완료 목록
    st.subheader("✅ 발주 완료 목록")

    if df_orders is not None and len(df_orders) > 0:
        # 최신 순으로 정렬
        df_display = df_orders.copy()
        df_display = df_display.sort_values('발주일', ascending=False)

        # 컬럼 선택 및 포맷
        display_cols = ['발주일', 'SKU코드', '제품명', 'ABC/XYZ', '현재고',
                        '발주량', '구매원가', '예상입고일']
        df_display = df_display[display_cols]

        st.dataframe(df_display, use_container_width=True, height=500)

        # Excel 다운로드
        from io import BytesIO
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_display.to_excel(writer, index=False, sheet_name='발주내역')

        output.seek(0)
        st.download_button(
            label="📥 발주 내역 Excel 다운로드",
            data=output,
            file_name=f"발주내역_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("📋 발주 내역이 없습니다. '자동 발주' 또는 '발주 관리' 탭에서 발주를 실행하세요.")

def show_simulation(df_analysis, df_psi):
    """시뮬레이션 및 예측 화면"""
    st.header("🎲 예측 및 시뮬레이션")

    st.markdown("""
    이 탭에서는 다양한 시나리오를 시뮬레이션하고 수요를 예측하여
    최적의 재고 관리 전략을 수립할 수 있습니다.
    """)

    # 세션 상태에 선택된 탭 저장
    if 'sim_selected_tab' not in st.session_state:
        st.session_state.sim_selected_tab = "📈 판매 추이 분석"

    # 탭 선택 콜백 함수
    def on_tab_change():
        st.session_state.sim_selected_tab = st.session_state.sim_tab_selector

    # 탭 선택 (radio 버튼 사용)
    sim_tab_option = st.radio(
        "분석 유형 선택:",
        ["📈 판매 추이 분석", "🎲 시나리오 분석", "📉 재고 최적화"],
        index=["📈 판매 추이 분석", "🎲 시나리오 분석", "📉 재고 최적화"].index(st.session_state.sim_selected_tab),
        horizontal=True,
        key='sim_tab_selector',
        on_change=on_tab_change
    )

    # 현재 선택된 탭 사용
    sim_tab_option = st.session_state.sim_selected_tab

    st.markdown("---")

    # ===== 1. 수요 예측 =====
    if sim_tab_option == "📈 판매 추이 분석":
        st.subheader("📈 판매 추이 분석 모델")
        st.markdown("이동평균과 지수평활법을 이용한 수요 예측")

        # 제품 선택
        if len(df_analysis) > 0:
            col1, col2 = st.columns([2, 1])

            with col1:
                selected_sku = st.selectbox(
                    "분석할 제품 선택:",
                    options=df_analysis['SKU코드'].tolist(),
                    format_func=lambda x: f"{x} - {df_analysis[df_analysis['SKU코드']==x]['제품명'].iloc[0]}"
                )

            with col2:
                forecast_days = st.number_input("예측 기간 (일)", min_value=7, max_value=90, value=30, step=7)

            if selected_sku:
                product_data = df_analysis[df_analysis['SKU코드'] == selected_sku].iloc[0]

                st.markdown("---")

                # 현재 상태
                col_info1, col_info2, col_info3 = st.columns(3)
                with col_info1:
                    st.metric("일평균 판매", f"{product_data['일평균판매']:.1f}개")
                with col_info2:
                    st.metric("현재고", f"{product_data['현재고']:,.0f}개")
                with col_info3:
                    계절성 = product_data.get('계절', '정보없음') if '계절' in product_data else '정보없음'
                    st.metric("계절성", 계절성)

                # 예측 모델
                st.markdown("### 예측 결과")

                daily_sales = product_data['일평균판매']

                # 1. 단순 이동평균 (7일)
                ma_forecast = daily_sales

                # 2. 지수평활법 (alpha=0.3)
                alpha = 0.3
                es_forecast = daily_sales

                # 3. 계절성 보정 (계절 정보 활용)
                seasonal_factor = 1.0
                if '계절' in product_data:
                    season = str(product_data['계절']).strip()
                    current_month = datetime.now().month
                    # 간단한 계절성 보정
                    if season == '여름' and current_month in [6, 7, 8]:
                        seasonal_factor = 1.2
                    elif season == '겨울' and current_month in [12, 1, 2]:
                        seasonal_factor = 1.2
                    elif season == '봄' and current_month in [3, 4, 5]:
                        seasonal_factor = 1.1
                    elif season == '가을' and current_month in [9, 10, 11]:
                        seasonal_factor = 1.1

                seasonal_forecast = daily_sales * seasonal_factor

                # 예측 데이터프레임 생성
                forecast_df = pd.DataFrame({
                    '일차': range(1, forecast_days + 1),
                    '단순평균': [ma_forecast] * forecast_days,
                    '지수평활': [es_forecast] * forecast_days,
                    '계절성보정': [seasonal_forecast] * forecast_days
                })

                # 그래프
                fig = go.Figure()
                fig.add_trace(go.Scatter(
                    x=forecast_df['일차'], y=forecast_df['단순평균'],
                    name='단순 이동평균', line=dict(color='#2196F3', width=2)
                ))
                fig.add_trace(go.Scatter(
                    x=forecast_df['일차'], y=forecast_df['지수평활'],
                    name='지수평활법', line=dict(color='#4CAF50', width=2, dash='dash')
                ))
                fig.add_trace(go.Scatter(
                    x=forecast_df['일차'], y=forecast_df['계절성보정'],
                    name='계절성 보정', line=dict(color='#FF9800', width=2, dash='dot')
                ))

                fig.update_layout(
                    title=f"{forecast_days}일 수요 예측",
                    xaxis_title="일차",
                    yaxis_title="예상 일판매량 (개)",
                    hovermode='x unified',
                    height=400
                )

                st.plotly_chart(fig, use_container_width=True)

                # 예측 요약
                st.markdown("### 예측 요약")
                col_sum1, col_sum2, col_sum3 = st.columns(3)

                with col_sum1:
                    st.info(f"""
                    **단순 이동평균**
                    - 일평균: {ma_forecast:.1f}개
                    - {forecast_days}일 총량: {ma_forecast * forecast_days:.0f}개
                    """)

                with col_sum2:
                    st.success(f"""
                    **지수평활법**
                    - 일평균: {es_forecast:.1f}개
                    - {forecast_days}일 총량: {es_forecast * forecast_days:.0f}개
                    """)

                with col_sum3:
                    seasonal_color = "warning" if seasonal_factor > 1 else "info"
                    st.markdown(f"""
                    <div style='background-color: #fff3cd; padding: 1rem; border-radius: 0.5rem; border-left: 4px solid #ff9800;'>
                    <strong>계절성 보정</strong><br>
                    - 보정계수: {seasonal_factor}x<br>
                    - 일평균: {seasonal_forecast:.1f}개<br>
                    - {forecast_days}일 총량: {seasonal_forecast * forecast_days:.0f}개
                    </div>
                    """, unsafe_allow_html=True)

        else:
            st.warning("⚠️ 분석할 데이터가 없습니다.")

    # ===== 2. 시나리오 분석 =====
    elif sim_tab_option == "🎲 시나리오 분석":
        st.subheader("🎲 시나리오 분석")
        st.markdown("다양한 상황 변화에 따른 재고 영향 분석")

        # 시나리오 선택
        scenario_type = st.radio(
            "시나리오 유형:",
            ["수요 변화", "리드타임 변화", "복합 시나리오"],
            horizontal=True
        )

        if scenario_type == "수요 변화":
            st.markdown("### 📊 수요 변화 시뮬레이션")

            col1, col2 = st.columns(2)
            with col1:
                demand_change = st.slider(
                    "수요 변화율 (%)",
                    min_value=-50,
                    max_value=100,
                    value=20,
                    step=5
                )

            with col2:
                apply_to = st.selectbox(
                    "적용 대상:",
                    ["전체 품목", "A등급만", "B등급만", "C등급만"]
                )

            # 필터 적용
            if apply_to == "전체 품목":
                sim_df = df_analysis.copy()
            else:
                grade = apply_to[0]  # 'A', 'B', 'C'
                sim_df = df_analysis[df_analysis['ABC등급'] == grade].copy()

            if len(sim_df) > 0:
                # 시뮬레이션 계산
                sim_df['시뮬_일평균판매'] = sim_df['일평균판매'] * (1 + demand_change / 100)
                sim_df['시뮬_발주점'] = (sim_df['시뮬_일평균판매'] * sim_df['리드타임']) + sim_df['안전재고']
                sim_df['시뮬_발주필요'] = sim_df['현재고'] <= sim_df['시뮬_발주점']
                sim_df['시뮬_재고소진일'] = sim_df.apply(
                    lambda row: int(row['현재고'] / row['시뮬_일평균판매']) if row['시뮬_일평균판매'] > 0 else 999,
                    axis=1
                )

                # 결과 비교
                st.markdown("### 시뮬레이션 결과")

                col_before, col_arrow, col_after = st.columns([2, 1, 2])

                with col_before:
                    st.markdown("**현재 상태**")
                    st.metric("발주 필요 품목", f"{len(sim_df[sim_df['발주필요']==True])}개")
                    st.metric("재고 위험 품목", f"{len(sim_df[sim_df['재고소진일']<=7])}개")
                    st.metric("평균 재고소진일", f"{sim_df[sim_df['재고소진일']<999]['재고소진일'].mean():.1f}일")

                with col_arrow:
                    st.markdown("<br><br>", unsafe_allow_html=True)
                    st.markdown(f"<div style='text-align: center; font-size: 3rem;'>→</div>", unsafe_allow_html=True)
                    st.markdown(f"<div style='text-align: center; font-weight: bold; color: {'red' if demand_change > 0 else 'blue'};'>{demand_change:+d}%</div>", unsafe_allow_html=True)

                with col_after:
                    st.markdown(f"**수요 {demand_change:+d}% 시**")
                    new_order_needed = len(sim_df[sim_df['시뮬_발주필요']==True])
                    new_risk = len(sim_df[sim_df['시뮬_재고소진일']<=7])
                    new_avg_days = sim_df[sim_df['시뮬_재고소진일']<999]['시뮬_재고소진일'].mean()

                    st.metric("발주 필요 품목", f"{new_order_needed}개",
                             delta=f"{new_order_needed - len(sim_df[sim_df['발주필요']==True]):+d}개")
                    st.metric("재고 위험 품목", f"{new_risk}개",
                             delta=f"{new_risk - len(sim_df[sim_df['재고소진일']<=7]):+d}개")
                    st.metric("평균 재고소진일", f"{new_avg_days:.1f}일",
                             delta=f"{new_avg_days - sim_df[sim_df['재고소진일']<999]['재고소진일'].mean():.1f}일")

                # 상세 리스트
                with st.expander("📋 영향받는 품목 상세 보기"):
                    # 새로 발주 필요해진 품목
                    new_orders = sim_df[(sim_df['시뮬_발주필요']==True) & (sim_df['발주필요']==False)]
                    if len(new_orders) > 0:
                        st.markdown(f"**새로 발주 필요해진 품목: {len(new_orders)}개**")
                        st.dataframe(new_orders[['SKU코드', '제품명', '현재고', '시뮬_발주점', '시뮬_재고소진일']],
                                   use_container_width=True)

        elif scenario_type == "리드타임 변화":
            st.markdown("### ⏱️ 리드타임 변화 시뮬레이션")

            col1, col2 = st.columns(2)
            with col1:
                leadtime_multiplier = st.slider(
                    "리드타임 배수",
                    min_value=0.5,
                    max_value=3.0,
                    value=2.0,
                    step=0.5
                )

            with col2:
                apply_to_lt = st.selectbox(
                    "적용 대상:",
                    ["전체 품목", "A등급만", "B등급만", "C등급만"],
                    key="lt_apply"
                )

            # 필터 적용
            if apply_to_lt == "전체 품목":
                sim_df_lt = df_analysis.copy()
            else:
                grade_lt = apply_to_lt[0]
                sim_df_lt = df_analysis[df_analysis['ABC등급'] == grade_lt].copy()

            if len(sim_df_lt) > 0:
                # 시뮬레이션
                sim_df_lt['시뮬_리드타임'] = sim_df_lt['리드타임'] * leadtime_multiplier
                sim_df_lt['시뮬_발주점'] = (sim_df_lt['일평균판매'] * sim_df_lt['시뮬_리드타임']) + sim_df_lt['안전재고']
                sim_df_lt['시뮬_부족량'] = (sim_df_lt['시뮬_발주점'] - sim_df_lt['현재고']).clip(lower=0)
                sim_df_lt['시뮬_권장발주량'] = sim_df_lt.apply(
                    lambda row: int(row['시뮬_부족량'] + row['일평균판매'] * 7) if row['시뮬_부족량'] > 0 else 0,
                    axis=1
                )

                # 결과
                st.markdown("### 시뮬레이션 결과")

                col_res1, col_res2, col_res3 = st.columns(3)

                with col_res1:
                    avg_lt_before = sim_df_lt['리드타임'].mean()
                    avg_lt_after = sim_df_lt['시뮬_리드타임'].mean()
                    st.metric("평균 리드타임",
                             f"{avg_lt_after:.1f}일",
                             delta=f"{avg_lt_after - avg_lt_before:.1f}일")

                with col_res2:
                    total_order_before = sim_df_lt['권장발주량'].sum()
                    total_order_after = sim_df_lt['시뮬_권장발주량'].sum()
                    st.metric("총 발주량",
                             f"{total_order_after:,.0f}개",
                             delta=f"{total_order_after - total_order_before:+,.0f}개")

                with col_res3:
                    items_need_order = len(sim_df_lt[sim_df_lt['시뮬_권장발주량'] > 0])
                    st.metric("발주 필요 품목", f"{items_need_order}개")

                # 그래프
                comparison_df = pd.DataFrame({
                    '품목': ['현재 발주량', f'리드타임 {leadtime_multiplier}x 시'],
                    '총량': [total_order_before, total_order_after]
                })

                fig_lt = px.bar(comparison_df, x='품목', y='총량',
                               title='총 발주량 비교',
                               color='품목',
                               color_discrete_map={'현재 발주량': '#2196F3', f'리드타임 {leadtime_multiplier}x 시': '#FF5722'})
                fig_lt.update_layout(showlegend=False, height=300)
                st.plotly_chart(fig_lt, use_container_width=True)

        else:  # 복합 시나리오
            st.markdown("### 🎯 복합 시나리오 시뮬레이션")
            st.markdown("수요와 리드타임이 동시에 변화할 때의 영향")

            col1, col2 = st.columns(2)
            with col1:
                combined_demand = st.slider("수요 변화율 (%)", -50, 100, 20, 5, key="combined_demand")
            with col2:
                combined_lt = st.slider("리드타임 배수", 0.5, 3.0, 1.5, 0.5, key="combined_lt")

            sim_df_combined = df_analysis.copy()

            # 복합 계산
            sim_df_combined['시뮬_일평균판매'] = sim_df_combined['일평균판매'] * (1 + combined_demand / 100)
            sim_df_combined['시뮬_리드타임'] = sim_df_combined['리드타임'] * combined_lt
            sim_df_combined['시뮬_발주점'] = (sim_df_combined['시뮬_일평균판매'] * sim_df_combined['시뮬_리드타임']) + sim_df_combined['안전재고']
            sim_df_combined['시뮬_부족량'] = (sim_df_combined['시뮬_발주점'] - sim_df_combined['현재고']).clip(lower=0)
            sim_df_combined['시뮬_권장발주량'] = sim_df_combined.apply(
                lambda row: int(row['시뮬_부족량'] + row['시뮬_일평균판매'] * 7) if row['시뮬_부족량'] > 0 else 0,
                axis=1
            )

            # 결과 대시보드
            st.markdown("### 📊 종합 영향 분석")

            metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)

            with metric_col1:
                current_total = df_analysis['권장발주량'].sum()
                sim_total = sim_df_combined['시뮬_권장발주량'].sum()
                st.metric("총 발주량", f"{sim_total:,.0f}개",
                         delta=f"{sim_total - current_total:+,.0f}개")

            with metric_col2:
                current_items = len(df_analysis[df_analysis['권장발주량'] > 0])
                sim_items = len(sim_df_combined[sim_df_combined['시뮬_권장발주량'] > 0])
                st.metric("발주 품목 수", f"{sim_items}개",
                         delta=f"{sim_items - current_items:+d}개")

            with metric_col3:
                current_risk = len(df_analysis[df_analysis['재고소진일'] <= 7])
                sim_df_combined['시뮬_재고소진일'] = sim_df_combined.apply(
                    lambda row: int(row['현재고'] / row['시뮬_일평균판매']) if row['시뮬_일평균판매'] > 0 else 999,
                    axis=1
                )
                sim_risk = len(sim_df_combined[sim_df_combined['시뮬_재고소진일'] <= 7])
                st.metric("재고 위험 품목", f"{sim_risk}개",
                         delta=f"{sim_risk - current_risk:+d}개")

            with metric_col4:
                avg_rop_before = df_analysis['발주점'].mean()
                avg_rop_after = sim_df_combined['시뮬_발주점'].mean()
                st.metric("평균 발주점", f"{avg_rop_after:.0f}개",
                         delta=f"{avg_rop_after - avg_rop_before:+.0f}개")

    # ===== 3. 재고 최적화 =====
    elif sim_tab_option == "📉 재고 최적화":
        st.subheader("📉 재고 최적화 시뮬레이터")
        st.markdown("목표 재고회전율 달성을 위한 최적 재고 수준 계산")

        col1, col2 = st.columns([1, 1])

        with col1:
            target_turnover = st.number_input(
                "목표 재고회전율 (회/년)",
                min_value=1.0,
                max_value=20.0,
                value=6.0,
                step=0.5,
                help="일반적으로 4-12회가 적정 범위입니다",
                key="target_turnover_input"
            )

        with col2:
            # 현재 재고회전율 계산 (NaN 처리)
            avg_days = df_analysis[df_analysis['재고소진일'] < 999]['재고소진일'].mean()
            if pd.isna(avg_days) or avg_days == 0:
                current_turnover = 0
            else:
                current_turnover = 365 / avg_days
            st.metric("현재 재고회전율", f"{current_turnover:.1f}회/년")

        if st.button("🎯 최적화 계산 실행", type="primary"):
            # 목표 회전율 달성을 위한 최적 재고 계산
            opt_df = df_analysis.copy()

            # 목표 재고소진일 = 365 / 목표회전율
            target_days = 365 / target_turnover

            # 최적 재고 = 일평균판매 × 목표재고소진일
            opt_df['최적_재고수준'] = (opt_df['일평균판매'] * target_days).astype(int)
            opt_df['재고_조정필요'] = opt_df['현재고'] - opt_df['최적_재고수준']
            opt_df['조정_방향'] = opt_df['재고_조정필요'].apply(
                lambda x: '🔴 감축' if x > 0 else ('🟢 증가' if x < 0 else '✅ 적정')
            )

            # 결과 요약
            st.markdown("### 최적화 결과")

            result_col1, result_col2, result_col3, result_col4 = st.columns(4)

            with result_col1:
                current_total_inv = opt_df['현재고'].sum()
                optimal_total_inv = opt_df['최적_재고수준'].sum()
                st.metric("총 재고",
                         f"{optimal_total_inv:,.0f}개",
                         delta=f"{optimal_total_inv - current_total_inv:+,.0f}개")

            with result_col2:
                reduce_items = len(opt_df[opt_df['재고_조정필요'] > 0])
                st.metric("감축 필요 품목", f"{reduce_items}개")

            with result_col3:
                increase_items = len(opt_df[opt_df['재고_조정필요'] < 0])
                st.metric("증가 필요 품목", f"{increase_items}개")

            with result_col4:
                optimal_items = len(opt_df[opt_df['재고_조정필요'] == 0])
                st.metric("적정 재고 품목", f"{optimal_items}개")

            # ABC등급별 조정 현황
            st.markdown("### ABC등급별 재고 조정")

            abc_opt = opt_df.groupby('ABC등급').agg({
                '현재고': 'sum',
                '최적_재고수준': 'sum',
                '재고_조정필요': 'sum'
            }).reset_index()
            abc_opt.columns = ['ABC등급', '현재고', '최적재고', '조정필요']

            fig_opt = go.Figure()
            fig_opt.add_trace(go.Bar(name='현재고', x=abc_opt['ABC등급'], y=abc_opt['현재고'], marker_color='#2196F3'))
            fig_opt.add_trace(go.Bar(name='최적재고', x=abc_opt['ABC등급'], y=abc_opt['최적재고'], marker_color='#4CAF50'))

            fig_opt.update_layout(
                barmode='group',
                title='ABC등급별 재고 현황 vs 최적 재고',
                xaxis_title='ABC등급',
                yaxis_title='재고량 (개)',
                height=400
            )

            st.plotly_chart(fig_opt, use_container_width=True)

            # 조정 필요 품목 리스트
            st.markdown("### 🔍 재고 조정 필요 품목 (상위 20개)")

            # 조정 폭이 큰 순서로 정렬
            top_adjustments = opt_df.nlargest(20, '재고_조정필요', keep='all')[
                ['SKU코드', '제품명', 'ABC등급', '현재고', '최적_재고수준', '재고_조정필요', '조정_방향']
            ]

            # 색상 스타일 함수 (진한 색상 + 흰색 텍스트)
            def color_adjustment(val):
                if isinstance(val, str):
                    if '감축' in val:
                        return 'background-color: #ef5350; color: white; font-weight: bold'
                    elif '증가' in val:
                        return 'background-color: #66bb6a; color: white; font-weight: bold'
                    elif '적정' in val:
                        return 'background-color: #42a5f5; color: white; font-weight: bold'
                return ''

            st.dataframe(
                top_adjustments.style.applymap(color_adjustment, subset=['조정_방향']),
                use_container_width=True
            )

            # Excel 다운로드 (데이터가 있을 때만)
            if len(opt_df) > 0:
                from io import BytesIO
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    opt_df[['SKU코드', '제품명', 'ABC등급', 'XYZ등급', '현재고', '최적_재고수준',
                           '재고_조정필요', '조정_방향']].to_excel(writer, sheet_name='최적화결과', index=False)
                    abc_opt.to_excel(writer, sheet_name='ABC별요약', index=False)

                st.download_button(
                    label="📥 최적화 결과 다운로드",
                    data=buffer.getvalue(),
                    file_name=f"재고최적화_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

def show_ai_assistant(dashboard_data, df_analysis):
    """AI 어시스턴트 챗봇"""
    st.header("🤖 AI 발주 어시스턴트")

    st.markdown("""
    재고 관리와 발주에 대해 무엇이든 물어보세요!

    **질문 예시:**
    - "긴급 발주 필요한 품목 알려줘"
    - "A등급 중 부족한 거 뭐야?"
    - "이번 주 발주 필요한 품목은?"
    - "재고 회전율이 뭐야?"
    """)

    # 채팅 히스토리 초기화
    if "chat_history" not in st.session_state:
        st.session_state.chat_history = []

    # 채팅 히스토리 표시
    for message in st.session_state.chat_history:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])

    # 사용자 입력
    if prompt := st.chat_input("질문을 입력하세요..."):
        # 사용자 메시지 추가
        st.session_state.chat_history.append({"role": "user", "content": prompt})
        with st.chat_message("user"):
            st.markdown(prompt)

        # AI 답변 생성
        with st.chat_message("assistant"):
            response = generate_ai_response(prompt, dashboard_data, df_analysis)
            st.markdown(response)

        # AI 답변 추가
        st.session_state.chat_history.append({"role": "assistant", "content": response})

def generate_ai_response(question, dashboard_data, df_analysis):
    """질문에 대한 AI 답변 생성 (Claude API 또는 규칙 기반)"""

    # Claude API 사용 가능 여부 확인
    api_key = st.session_state.get('claude_api_key', '')

    if ANTHROPIC_AVAILABLE and api_key:
        # Claude API 사용
        try:
            return generate_ai_response_with_api(question, dashboard_data, df_analysis, api_key)
        except Exception as e:
            st.error(f"⚠️ API 호출 실패: {str(e)}")
            # API 실패 시 규칙 기반으로 폴백
            return generate_rule_based_response(question, dashboard_data, df_analysis)
    else:
        # 규칙 기반 챗봇
        return generate_rule_based_response(question, dashboard_data, df_analysis)

def generate_ai_response_with_api(question, dashboard_data, df_analysis, api_key):
    """Claude API를 사용한 실제 AI 답변 생성"""

    # 재고 데이터 요약
    context = f"""
현재 재고 현황:
- 총 SKU: {dashboard_data['total_sku']}개
- 총 재고금액: {dashboard_data['total_value']/100000000:.1f}억원
- 평균 재고소진일: {dashboard_data['avg_turnover_days']:.1f}일

재고 상태별:
- 🔴 부족: {len(df_analysis[df_analysis['재고상태']=='🔴 부족'])}개
- 🟡 재주문 필요: {len(df_analysis[df_analysis['재고상태']=='🟡 재주문 필요'])}개
- 🟢 적정: {len(df_analysis[df_analysis['재고상태']=='🟢 적정'])}개
- 🔵 과잉: {len(df_analysis[df_analysis['재고상태']=='🔵 과잉'])}개

긴급 발주 필요 품목:
"""
    # 긴급 발주 품목 추가
    urgent = df_analysis[df_analysis['재고상태'] == '🔴 부족'].head(10)
    if len(urgent) > 0:
        for idx, row in urgent.iterrows():
            context += f"\n- {row['SKU코드']} ({row['제품명']}): 현재고 {row['현재고']:.0f}개, ABC등급 {row['ABC등급']}, XYZ등급 {row.get('XYZ등급', 'N/A')}"
    else:
        context += "\n(없음)"

    # 발주 필요 품목 추가
    need_order = df_analysis[df_analysis['발주필요'] == True].head(10)
    if len(need_order) > 0:
        context += "\n\n발주 필요 품목:\n"
        for idx, row in need_order.iterrows():
            context += f"- {row['SKU코드']}: 권장 {row['권장발주량']:.0f}개\n"

    # Claude API 호출
    client = anthropic.Anthropic(api_key=api_key)

    message = client.messages.create(
        model="claude-3-5-sonnet-20241022",
        max_tokens=1024,
        system="""당신은 재고 관리 및 발주 전문가입니다.
사용자의 질문에 대해 정확하고 실무적인 답변을 제공하세요.
- 간결하고 명확하게 답변하세요
- 구체적인 수치와 품목 정보를 제공하세요
- 실무에 도움이 되는 조언을 제공하세요
- 한국어로 답변하세요""",
        messages=[
            {
                "role": "user",
                "content": f"{context}\n\n사용자 질문: {question}"
            }
        ]
    )

    return message.content[0].text

def generate_rule_based_response(question, dashboard_data, df_analysis):
    """규칙 기반 챗봇 답변 (30+ 패턴)"""
    question_lower = question.lower()

    # ==================== 재고 상태 조회 ====================

    # 1. 긴급 발주 / 부족 품목
    if any(keyword in question_lower for keyword in ['긴급', '부족', '급한', '위험', '품절']):
        urgent = df_analysis[df_analysis['재고상태'] == '🔴 부족']
        if len(urgent) > 0:
            items = urgent.head(10)
            response = f"**🚨 긴급 발주 필요 품목 {len(urgent)}개:**\n\n"
            for idx, row in items.iterrows():
                days = row.get('재고소진일', 0)
                response += f"- **{row['SKU코드']}** ({row['제품명']})\n"
                response += f"  현재고: {row['현재고']:.0f}개 | 소진일: {days:.0f}일 | ABC: {row['ABC등급']}\n"
            return response
        else:
            return "✅ 현재 긴급 발주가 필요한 품목이 없습니다!"

    # 2. 과잉 재고 품목
    elif any(keyword in question_lower for keyword in ['과잉', '많은', '재고많', '넘치']):
        excess = df_analysis[df_analysis['재고상태'] == '🔵 과잉']
        if len(excess) > 0:
            items = excess.head(10)
            response = f"**🔵 과잉 재고 품목 {len(excess)}개:**\n\n"
            for idx, row in items.iterrows():
                response += f"- **{row['SKU코드']}** ({row['제품명']})\n"
                response += f"  현재고: {row['현재고']:.0f}개 | 소진일: {row.get('재고소진일', 0):.0f}일\n"
            return response
        else:
            return "✅ 과잉 재고 품목이 없습니다."

    # 3. 적정 재고 품목
    elif '적정' in question_lower and '재고' in question_lower:
        optimal = df_analysis[df_analysis['재고상태'] == '🟢 적정']
        return f"**🟢 적정 재고 품목: {len(optimal)}개**\n\n재고 관리가 잘 되고 있는 품목입니다!"

    # ==================== ABC 등급별 조회 ====================

    # 4. A등급 품목
    elif 'a등급' in question_lower or 'a급' in question_lower:
        a_items = df_analysis[df_analysis['ABC등급'] == 'A']
        if '부족' in question_lower:
            shortage = a_items[a_items['재고상태'] == '🔴 부족']
            if len(shortage) > 0:
                response = f"**🚨 A등급 중 부족 품목 {len(shortage)}개:**\n\n"
                for idx, row in shortage.head(5).iterrows():
                    response += f"- **{row['SKU코드']}**: 현재고 {row['현재고']:.0f}개, 발주점 {row['발주점']:.0f}개\n"
                return response
            else:
                return "✅ A등급 품목 중 부족한 품목이 없습니다!"
        else:
            shortage_cnt = len(a_items[a_items['재고상태']=='🔴 부족'])
            optimal_cnt = len(a_items[a_items['재고상태']=='🟢 적정'])
            excess_cnt = len(a_items[a_items['재고상태']=='🔵 과잉'])
            return f"""**📊 A등급 품목 현황: 총 {len(a_items)}개**

- 🔴 부족: {shortage_cnt}개
- 🟢 적정: {optimal_cnt}개
- 🔵 과잉: {excess_cnt}개

**매출 상위 품목**이므로 철저한 관리가 필요합니다!"""

    # 5. B등급 품목
    elif 'b등급' in question_lower or 'b급' in question_lower:
        b_items = df_analysis[df_analysis['ABC등급'] == 'B']
        shortage_cnt = len(b_items[b_items['재고상태']=='🔴 부족'])
        optimal_cnt = len(b_items[b_items['재고상태']=='🟢 적정'])
        return f"""**📊 B등급 품목 현황: 총 {len(b_items)}개**

- 🔴 부족: {shortage_cnt}개
- 🟢 적정: {optimal_cnt}개

**중요도 중간** 품목으로 효율적 관리가 필요합니다."""

    # 6. C등급 품목
    elif 'c등급' in question_lower or 'c급' in question_lower:
        c_items = df_analysis[df_analysis['ABC등급'] == 'C']
        excess_cnt = len(c_items[c_items['재고상태']=='🔵 과잉'])
        return f"""**📊 C등급 품목 현황: 총 {len(c_items)}개**

- 🔵 과잉: {excess_cnt}개

**저회전 품목**이므로 과잉 재고를 줄이는 것이 중요합니다."""

    # ==================== 발주 관련 ====================

    # 7. 발주 필요 품목
    elif '발주' in question_lower and any(kw in question_lower for kw in ['필요', '해야', '할거', '목록', '리스트']):
        need_order = df_analysis[df_analysis['발주필요'] == True]
        if len(need_order) > 0:
            # ABC 우선순위로 정렬
            need_order_sorted = need_order.sort_values('ABC등급')

            response = f"**📦 발주 필요 품목 {len(need_order)}개:**\n\n"
            response += "**[A등급 우선]**\n"

            for idx, row in need_order_sorted.head(10).iterrows():
                moq = row.get('MOQ', 0)
                response += f"- **{row['SKU코드']}** ({row['ABC등급']}등급)\n"
                response += f"  권장: {row['권장발주량']:.0f}개 | MOQ: {moq:.0f}개\n"

            if len(need_order) > 10:
                response += f"\n...외 {len(need_order)-10}개 품목"
            return response
        else:
            return "✅ 현재 발주가 필요한 품목이 없습니다!"

    # 8. 발주 금액
    elif '발주' in question_lower and any(kw in question_lower for kw in ['금액', '비용', '예산', '얼마']):
        need_order = df_analysis[df_analysis['발주필요'] == True]
        if len(need_order) > 0 and '매입원가' in need_order.columns:
            total_cost = (need_order['권장발주량'] * need_order['매입원가']).sum()
            a_cost = (need_order[need_order['ABC등급']=='A']['권장발주량'] * need_order[need_order['ABC등급']=='A']['매입원가']).sum()
            return f"""**💰 발주 예상 금액:**

- 발주 품목: {len(need_order)}개
- 총 예상 금액: **{total_cost/10000:.0f}만원** ({total_cost/100000000:.1f}억원)

**ABC별 금액:**
- A등급: {a_cost/10000:.0f}만원"""
        else:
            return "발주 필요 품목이 없거나 매입원가 데이터가 없습니다."

    # ==================== 재고 분석 ====================

    # 9. 재고 회전율
    elif '재고회전' in question_lower or '회전율' in question_lower:
        if '개선' in question_lower or '높이' in question_lower:
            return """**📈 재고 회전율 개선 방법:**

**1. 느린 회전 품목 재고 감소**
   - C등급 과잉 재고 정리
   - 안전재고 수준 조정

**2. 발주 주기 최적화**
   - A등급: 자주, 소량 발주
   - C등급: 긴 주기, 필요시만 발주

**3. ABC/XYZ 분석 활용**
   - AX: 높은 회전율 유지 (자주 체크)
   - CZ: 최소 재고 유지

**4. 정기적 재고 점검**
   - 주간 단위 재고 현황 확인
   - 휴면 재고 발견 즉시 처리"""
        else:
            return f"""**📊 재고 회전율 정보:**

**현재 평균 재고소진일:** {dashboard_data['avg_turnover_days']:.1f}일
**연간 예상 회전율:** {365/dashboard_data['avg_turnover_days']:.1f}회/년

**재고회전율이란?**
재고가 1년에 몇 번 판매(회전)되는지를 나타내는 지표입니다.

**계산 방법:** 365 / 평균재고소진일

**기준:**
- 4~12회: 일반적인 범위
- 높을수록: 재고 효율 좋음
- 너무 높으면: 품절 위험 증가

💡 **분석 탭**에서 상세 정보를 확인하세요!"""

    # 10. 재고 현황 전체
    elif '재고' in question_lower and any(kw in question_lower for kw in ['현황', '상태', '요약', '전체']):
        shortage = len(df_analysis[df_analysis['재고상태']=='🔴 부족'])
        reorder = len(df_analysis[df_analysis['재고상태']=='🟡 재주문 필요'])
        optimal = len(df_analysis[df_analysis['재고상태']=='🟢 적정'])
        excess = len(df_analysis[df_analysis['재고상태']=='🔵 과잉'])

        return f"""**📦 전체 재고 현황 요약:**

**기본 정보:**
- 총 SKU: {dashboard_data['total_sku']}개
- 총 재고금액: {dashboard_data['total_value']/100000000:.1f}억원
- 평균 재고소진일: {dashboard_data['avg_turnover_days']:.1f}일

**재고 상태별:**
- 🔴 부족: {shortage}개 ({shortage/dashboard_data['total_sku']*100:.1f}%)
- 🟡 재주문 필요: {reorder}개
- 🟢 적정: {optimal}개
- 🔵 과잉: {excess}개

**조치 필요:**
- 긴급 발주: {shortage}개
- 발주 검토: {reorder}개
- 재고 감축 검토: {excess}개"""

    # ==================== 용어 설명 ====================

    # 11. ABC 분석이란
    elif 'abc' in question_lower and any(kw in question_lower for kw in ['뭐', '무엇', '설명', '란']):
        return """**📊 ABC 분석이란?**

**정의:**
매출 기여도에 따라 품목을 3등급으로 분류하는 방법

**등급 기준:**
- **A등급 (20%)**: 매출의 80% 차지 → **최우선 관리**
- **B등급 (30%)**: 매출의 15% 차지 → 중요도 중간
- **C등급 (50%)**: 매출의 5% 차지 → 효율적 관리

**실무 활용:**
- A등급: 자주 체크, 높은 안전재고, 우선 발주
- B등급: 정기 체크, 적정 안전재고
- C등급: 최소 재고, 장기 발주 주기

**파레토 법칙 (80:20 법칙)** 기반입니다!"""

    # 12. 안전재고란
    elif '안전재고' in question_lower and any(kw in question_lower for kw in ['뭐', '무엇', '설명', '란']):
        return """**🛡️ 안전재고란?**

**정의:**
예상치 못한 수요 증가나 공급 지연에 대비한 여유 재고

**왜 필요한가?**
- 수요 급증 대응
- 공급업체 지연 대비
- 품절 방지

**계산 방법:**
안전재고 = 일평균판매 × 리드타임 × 안전계수

**등급별 안전계수:**
- A등급: 1.5배 (높은 안전재고)
- B등급: 1.2배
- C등급: 1.0배 (최소 안전재고)

너무 높으면 → 재고비용 증가
너무 낮으면 → 품절 위험 증가"""

    # ==================== 시스템 사용법 ====================

    # 13. 발주 방법
    elif '발주' in question_lower and any(kw in question_lower for kw in ['방법', '어떻게', '하는법']):
        return """**📦 발주하는 방법:**

**1단계: 발주 관리 탭으로 이동**
   - 상단 메뉴에서 "📦 발주 관리" 클릭

**2단계: 필터 설정**
   - ABC 등급 선택 (A등급 우선)
   - 재고상태 선택 (부족/재주문 필요)

**3단계: 발주 실행**
   - 발주할 품목 선택
   - 권장발주량 확인 (MOQ 고려)
   - "✅ 발주 실행" 버튼 클릭

**4단계: 확인**
   - "📋 발주 현황" 탭에서 발주 내역 확인
   - PSI 파일에 자동 기록됨"""

    # 14. 이번 주 할 일
    elif any(kw in question_lower for kw in ['이번주', '이번 주', '오늘', '해야할']):
        need_order = df_analysis[df_analysis['발주필요'] == True]
        urgent = df_analysis[df_analysis['재고상태'] == '🔴 부족']

        return f"""**📅 이번 주 주요 업무:**

**1. 긴급 발주 ({len(urgent)}개)**
   - 부족 품목 즉시 발주
   - A등급 우선 처리

**2. 재주문 검토 ({len(need_order)}개)**
   - 발주점 이하 품목 확인
   - 발주 계획 수립

**3. 재고 점검**
   - 과잉 재고 확인
   - 휴면 재고 처리 방안

**4. 데이터 업데이트**
   - 최신 PSI 파일 업로드
   - KPI 지표 확인

💡 **발주 관리 탭**에서 바로 시작하세요!"""

    # 15. 도움말
    elif any(kw in question_lower for kw in ['도움말', '사용법', '어떻게', '모르겠', '처음']):
        return """**📖 스마트 발주 시스템 사용 가이드**

**1. 📊 대시보드**
   - 전체 재고 현황 한눈에 확인
   - KPI 지표, ABC/XYZ 차트

**2. 📦 발주 관리**
   - 발주 필요 품목 확인
   - 발주 실행 (권장수량 자동 계산)

**3. 📋 발주 현황**
   - 발주 내역 조회
   - Excel 다운로드

**4. 🤖 AI 어시스턴트**
   - 궁금한 점 질문
   - 실시간 데이터 분석

💡 **각 탭을 클릭해서 기능을 확인하세요!**"""

    # 16. 인사
    elif any(kw in question_lower for kw in ['안녕', '하이', '헬로', 'hi', 'hello']):
        return f"""안녕하세요! 👋 스마트 발주 시스템 AI 어시스턴트입니다.

**현재 재고 상황:**
- 총 SKU: {dashboard_data['total_sku']}개
- 부족 품목: {len(df_analysis[df_analysis['재고상태']=='🔴 부족'])}개
- 발주 필요: {len(df_analysis[df_analysis['발주필요']==True])}개

무엇을 도와드릴까요? 😊"""

    # 17. 감사 인사
    elif any(kw in question_lower for kw in ['고마워', '감사', '수고', '잘했어', '좋아']):
        return "😊 도움이 되었다니 기쁩니다! 다른 궁금한 점이 있으면 언제든 물어보세요!"

    # ==================== 기본 답변 ====================

    else:
        return """죄송합니다. 질문을 이해하지 못했습니다. 😅

**💬 자주 묻는 질문:**

**📦 재고 조회:**
- "긴급 발주 필요한 품목 알려줘"
- "과잉 재고 품목 보여줘"
- "재고 현황 요약해줘"

**📊 등급별 분석:**
- "A등급 품목 현황"
- "B등급 중 부족한 거"
- "ABC 분석이란?"

**🎯 발주 관련:**
- "발주 필요한 품목 목록"
- "발주 예상 금액 얼마야?"
- "발주하는 방법"

**📈 분석:**
- "재고 회전율 개선 방법"
- "안전재고란?"

**ℹ️ 시스템:**
- "이번 주 해야 할 일"
- "도움말"

**발주 관리 탭**에서 직접 확인하실 수도 있습니다!"""

def show_auto_orders(df_analysis):
    """자동 발주 시스템"""
    st.header("⚡ AI 자동 발주 시스템")

    st.markdown("""
    **🎯 수요 예측 기반 자동 발주 생성**
    - AI가 과거 판매 데이터를 분석하여 수요를 예측합니다
    - 동적 발주점을 계산하여 최적 발주 시점을 제안합니다
    - 우선순위 순으로 자동 정렬됩니다
    """)

    # 자동 발주 생성
    with st.spinner('🤖 AI가 수요를 분석하고 자동 발주를 생성하는 중...'):
        auto_orders = generate_auto_orders(df_analysis)

    if not auto_orders:
        st.success("✅ 현재 발주가 필요한 품목이 없습니다!")
        st.info("💡 모든 품목의 재고가 적정 수준입니다.")
        return

    # 통계 요약
    st.markdown("---")
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        high_priority = len([o for o in auto_orders if o['우선순위'] == 'HIGH'])
        st.metric("🔴 긴급 발주", f"{high_priority}개")

    with col2:
        medium_priority = len([o for o in auto_orders if o['우선순위'] == 'MEDIUM'])
        st.metric("🟡 일반 발주", f"{medium_priority}개")

    with col3:
        total_amount = sum([o['예상_발주금액'] for o in auto_orders])
        st.metric("💰 총 예상 금액", f"{total_amount/10000:.0f}만원")

    with col4:
        avg_confidence = np.mean([o['예측_신뢰도'] for o in auto_orders])
        st.metric("📊 평균 신뢰도", f"{avg_confidence:.0f}%")

    st.markdown("---")

    # 필터
    st.subheader("🔍 필터")
    filter_col1, filter_col2, filter_col3 = st.columns(3)

    with filter_col1:
        priority_filter = st.multiselect(
            "우선순위",
            options=['HIGH', 'MEDIUM', 'LOW'],
            default=['HIGH', 'MEDIUM']
        )

    with filter_col2:
        abc_filter = st.multiselect(
            "ABC 등급",
            options=['A', 'B', 'C'],
            default=['A', 'B', 'C']
        )

    with filter_col3:
        min_confidence = st.slider(
            "최소 신뢰도 (%)",
            0, 100, 60
        )

    # 필터 적용
    filtered_orders = [
        o for o in auto_orders
        if o['우선순위'] in priority_filter
        and o['ABC등급'] in abc_filter
        and o['예측_신뢰도'] >= min_confidence
    ]

    st.info(f"📋 {len(filtered_orders)}개 품목이 필터 조건을 만족합니다.")

    # 자동 발주 목록 표시
    st.subheader(f"📦 자동 발주 추천 목록 ({len(filtered_orders)}개)")

    if not filtered_orders:
        st.warning("⚠️ 필터 조건에 맞는 품목이 없습니다.")
        return

    # 선택된 품목 관리
    if 'auto_selected_items' not in st.session_state:
        st.session_state.auto_selected_items = set()

    # 전체 선택/해제 버튼
    col_sel1, col_sel2, col_sel3 = st.columns([1, 1, 3])
    with col_sel1:
        if st.button("✅ 전체 선택", key="auto_select_all"):
            st.session_state.auto_selected_items = set([o['SKU코드'] for o in filtered_orders])
    with col_sel2:
        if st.button("❌ 전체 해제", key="auto_deselect_all"):
            st.session_state.auto_selected_items = set()
    with col_sel3:
        selected_count = len(st.session_state.auto_selected_items)
        if selected_count > 0:
            st.success(f"✅ 선택된 품목: **{selected_count}개**")

    st.markdown("<br>", unsafe_allow_html=True)

    # 각 품목을 체크박스와 함께 표시
    for idx, order in enumerate(filtered_orders):
        sku_code = order['SKU코드']

        # 체크박스 초기값
        checkbox_key = f"auto_check_{sku_code}_{idx}"
        is_checked = sku_code in st.session_state.auto_selected_items
        if checkbox_key not in st.session_state:
            st.session_state[checkbox_key] = is_checked
        if is_checked or st.session_state.get(checkbox_key, False):
            st.session_state[checkbox_key] = True

        # 체크박스와 품목 정보
        col_check, col_info = st.columns([0.3, 4.7])

        with col_check:
            checked = st.checkbox("선택", key=checkbox_key, label_visibility="collapsed")
            if checked and sku_code not in st.session_state.auto_selected_items:
                st.session_state.auto_selected_items.add(sku_code)
            elif not checked and sku_code in st.session_state.auto_selected_items:
                st.session_state.auto_selected_items.remove(sku_code)

        with col_info:
            # 우선순위별 색상
            if order['우선순위'] == 'HIGH':
                priority_color = '🔴'
                bg_color = '#FEE2E2'
            elif order['우선순위'] == 'MEDIUM':
                priority_color = '🟡'
                bg_color = '#FEF3C7'
            else:
                priority_color = '⚪'
                bg_color = '#F3F4F6'

            with st.expander(
                f"{priority_color} **{order['SKU코드']}** - {order['제품명']} ({order['ABC등급']}등급)",
                expanded=False
            ):
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("현재고", f"{order['현재고']:,.0f}개")
                    st.metric("발주점", f"{order['발주점']:,.0f}개")
                with col2:
                    st.metric("권장발주량", f"{order['권장발주량']:,.0f}개")
                    st.metric("예측판매", f"{order['예측_일평균판매']:.1f}개/일")
                with col3:
                    st.metric("재고소진일", f"{order['재고소진일']:.0f}일")
                    st.metric("신뢰도", f"{order['예측_신뢰도']:.0f}%")
                with col4:
                    st.metric("예상금액", f"{order['예상_발주금액']/10000:.0f}만원")
                    st.metric("트렌드", order['예측_트렌드'])

                st.caption(f"XYZ등급: {order['XYZ등급']} | 리드타임: {order['리드타임']}일")

    # 일괄 승인 섹션
    st.markdown("---")
    st.subheader("✅ 발주 승인 및 실행")

    col_approve1, col_approve2 = st.columns([3, 1])

    with col_approve1:
        st.markdown("""
        **발주 승인 프로세스:**
        1. 위 목록에서 발주할 품목을 확인하세요
        2. 우선순위와 예상 금액을 검토하세요
        3. 아래 버튼으로 일괄 승인하거나, 발주 관리 탭에서 개별 승인하세요
        """)

    with col_approve2:
        selected_count = len(st.session_state.auto_selected_items)
        st.metric("선택된 품목", f"{selected_count}개")

    # 일괄 승인 버튼
    if st.button("⚡ 선택 품목 자동 발주 실행", type="primary", use_container_width=True):
        # 선택된 품목만 필터링
        selected_orders = [o for o in filtered_orders if o['SKU코드'] in st.session_state.auto_selected_items]

        if not selected_orders:
            st.error("❌ 발주할 품목을 먼저 선택하세요!")
        else:
            # 승인 확인
            total_amount = sum([o['예상_발주금액'] for o in selected_orders])
            total_qty = sum([o['권장발주량'] for o in selected_orders])

            st.warning(f"""
            ⚠️ **발주 확인**
            - 발주 품목: **{len(selected_orders)}개**
            - 총 발주량: **{total_qty:,.0f}개**
            - 예상 금액: **{total_amount/10000:,.0f}만원**

            아래 "최종 승인" 버튼을 눌러 발주를 실행하세요.
            """)

            # 최종 승인 버튼
            if st.button("✅ 최종 승인 및 발주 실행", type="secondary", use_container_width=True):
                # PSI 파일 경로 확인
                psi_file = st.session_state.get('psi_file_path', 'current_psi.xlsx')

                if not os.path.exists(psi_file):
                    st.error("❌ PSI 파일을 찾을 수 없습니다. 먼저 데이터를 업로드하세요.")
                else:
                    # 발주 실행
                    success_count = 0
                    fail_count = 0

                    with st.spinner('⚡ 자동 발주 실행 중...'):
                        for order in selected_orders:
                            try:
                                # 발주 데이터 준비
                                order_data = {
                                    'SKU코드': order['SKU코드'],
                                    '제품명': order['제품명'],
                                    'ABC등급': order['ABC등급'],
                                    'XYZ등급': order['XYZ등급'],
                                    '현재고': order['현재고'],
                                    '발주량': order['권장발주량'],
                                    '매입원가': order['매입원가'],
                                    '재고소진일': order['재고소진일'],
                                    '리드타임': order['리드타임'],
                                    '일평균판매': order['예측_일평균판매']
                                }

                                # 엑셀에 기록
                                record_order_to_excel(psi_file, order_data)
                                success_count += 1

                            except Exception as e:
                                fail_count += 1
                                st.error(f"❌ {order['SKU코드']} 발주 실패: {str(e)}")

                    # 결과 표시
                    st.markdown("---")
                    st.success(f"""
                    ✅ **자동 발주 완료!**

                    - ✅ 성공: **{success_count}개** 품목
                    - ❌ 실패: **{fail_count}개** 품목
                    - 📋 총 발주량: **{total_qty:,.0f}개**
                    - 💰 총 금액: **{total_amount/10000:,.0f}만원**

                    📋 **발주 현황 탭**에서 발주 내역을 확인하세요!
                    """)

                    if success_count > 0:
                        st.balloons()

                    # 발주 내역 저장 및 선택 초기화
                    if 'order_history' not in st.session_state:
                        st.session_state.order_history = {}

                    for order in selected_orders:
                        st.session_state.order_history[order['SKU코드']] = {
                            'quantity': order['권장발주량'],
                            'timestamp': datetime.now()
                        }

                    # 선택 초기화
                    st.session_state.auto_selected_items = set()

    # Excel 다운로드
    st.markdown("---")
    st.subheader("📥 자동 발주 목록 다운로드")

    # DataFrame 생성 (다운로드용)
    df_auto = pd.DataFrame(filtered_orders)

    # 데이터가 있을 때만 다운로드 버튼 표시
    if len(df_auto) > 0:
        from io import BytesIO
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_auto.to_excel(writer, sheet_name='자동발주추천', index=False)

        st.download_button(
            label="📥 Excel 다운로드",
            data=buffer.getvalue(),
            file_name=f"자동발주_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("📋 다운로드할 데이터가 없습니다.")

    # 수요 예측 상세
    st.markdown("---")
    with st.expander("🔍 수요 예측 상세 정보"):
        st.subheader("📊 예측 방법론")
        st.markdown("""
        **현재 사용 중인 예측 모델:**
        - **단순 평균법**: 과거 일평균 판매량 기반
        - **지수평활법**: 최근 데이터에 더 높은 가중치
        - **트렌드 분석**: 판매 추세 감지 (증가/감소/안정)

        **신뢰도 계산:**
        - 변동계수(CV) 기반
        - CV < 0.2: 신뢰도 90% (매우 안정)
        - CV < 0.5: 신뢰도 75% (보통)
        - CV < 1.0: 신뢰도 60% (변동 있음)
        - CV ≥ 1.0: 신뢰도 40% (불안정)

        **동적 발주점:**
        - 발주점 = 안전재고 + (예측 판매량 × 리드타임)
        - 안전재고 = 예측 판매량 × 리드타임 × 안전계수
        - 안전계수: A등급(1.5), B등급(1.2), C등급(1.0)
        """)

        st.subheader("🎯 우선순위 규칙")
        st.markdown("""
        **HIGH (긴급):**
        - 재고 상태가 '부족'인 품목
        - A등급 품목 우선

        **MEDIUM (일반):**
        - 재고 상태가 '재주문 필요'인 품목
        - B등급 품목

        **LOW (낮음):**
        - 그 외 발주 필요 품목
        - C등급 품목
        """)

def show_settings():
    """설정 화면"""
    st.header("⚙️ 시스템 설정")

    # AI API 키 설정 (Form 밖에서 따로 관리)
    st.subheader("🤖 AI 어시스턴트 설정")

    if ANTHROPIC_AVAILABLE:
        st.info("💡 Claude API를 연동하면 자연스러운 대화가 가능합니다!")

        # 세션 상태에 API 키 저장
        if 'claude_api_key' not in st.session_state:
            st.session_state.claude_api_key = ""

        api_key_input = st.text_input(
            "Claude API 키",
            value=st.session_state.claude_api_key,
            type="password",
            placeholder="sk-ant-api03-...",
            help="https://console.anthropic.com 에서 발급받으세요"
        )

        col_api1, col_api2 = st.columns([1, 3])
        with col_api1:
            if st.button("💾 API 키 저장", use_container_width=True):
                st.session_state.claude_api_key = api_key_input
                st.success("✅ API 키가 저장되었습니다!")

        with col_api2:
            if st.session_state.claude_api_key:
                st.success("🟢 API 키 설정됨 - AI 어시스턴트 사용 가능")
            else:
                st.warning("⚪ API 키 미설정 - 규칙 기반 챗봇으로 동작")
    else:
        st.warning("⚠️ anthropic 라이브러리가 설치되지 않았습니다.\n\n터미널에서 실행: `pip install anthropic`")

    st.markdown("---")

    st.subheader("발주 정책 설정")

    # Form으로 감싸서 저장 버튼을 누를 때만 적용
    with st.form("settings_form"):
        col1, col2 = st.columns(2)

        with col1:
            st.write("**A등급 설정**")
            a_safety_multiplier = st.slider("안전재고 배수", 1.0, 2.0, 1.5, 0.1, key="a_safety")
            a_order_cycle = st.selectbox("발주 주기", ["주 1회", "월 1회", "격주 1회"], index=1, key="a_cycle")

            st.write("**B등급 설정**")
            b_safety_multiplier = st.slider("안전재고 배수", 1.0, 2.0, 1.2, 0.1, key="b_safety")
            b_order_cycle = st.selectbox("발주 주기", ["월 1회", "격월 1회", "분기 1회"], index=1, key="b_cycle")

        with col2:
            st.write("**C등급 설정**")
            c_safety_multiplier = st.slider("안전재고 배수", 0.5, 1.5, 1.0, 0.1, key="c_safety")
            c_order_cycle = st.selectbox("발주 주기", ["분기 1회", "반기 1회", "수요 기반"], index=0, key="c_cycle")

            st.write("**알림 설정**")
            email_notification = st.checkbox("이메일 알림", value=True, key="email_notif")
            kakao_notification = st.checkbox("카카오톡 알림", value=False, key="kakao_notif")

        st.write("**리드타임 설정**")
        default_leadtime = st.number_input("기본 리드타임 (일)", min_value=1, value=30, step=1, key="default_lt")

        # 저장 버튼
        submitted = st.form_submit_button("💾 설정 저장", type="primary", use_container_width=True)

        if submitted:
            st.success("✅ 설정이 저장되었습니다!")

    st.markdown("---")

    st.markdown("---")

    st.subheader("📥 PSI 파일 관리")

    col_down, col_info = st.columns([1, 2])

    with col_down:
        # PSI 파일 다운로드
        if 'psi_file_path' in st.session_state:
            try:
                with open(st.session_state.psi_file_path, 'rb') as f:
                    psi_data = f.read()

                file_name = os.path.basename(st.session_state.psi_file_path)
                st.download_button(
                    label="📥 PSI 파일 다운로드",
                    data=psi_data,
                    file_name=f"PSI_updated_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    help="발주 기록이 포함된 최신 PSI 파일 다운로드"
                )
            except Exception as e:
                st.error(f"❌ 파일 다운로드 오류: {str(e)}")
        else:
            st.info("ℹ️ PSI 파일을 먼저 로드하세요")

    with col_info:
        st.info("""
        **📋 PSI 파일 다운로드/업로드**

        - **다운로드**: 발주 기록이 포함된 최신 PSI 파일
        - **업로드**: 좌측 사이드바에서 "파일 업로드" 선택
        - **발주리스트**: 자동으로 기록됨
        """)

    st.markdown("---")

    # 일일 리포트 다운로드
    st.subheader("📊 일일 리포트")

    col_report, col_info_report = st.columns([1, 2])

    with col_report:
        if st.button("📥 일일 리포트 다운로드", type="secondary", use_container_width=True):
            # 일일 리포트 Excel 생성
            if 'df_analysis' in st.session_state and st.session_state.df_analysis is not None:
                from io import BytesIO
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill

                wb = openpyxl.Workbook()

                # 시트 1: 요약
                ws_summary = wb.active
                ws_summary.title = "일일요약"

                ws_summary['A1'] = '📊 일일 재고 리포트'
                ws_summary['A1'].font = Font(size=16, bold=True)
                ws_summary['A2'] = f'생성일: {datetime.now().strftime("%Y-%m-%d %H:%M")}'

                ws_summary['A4'] = '항목'
                ws_summary['B4'] = '값'
                for cell in [ws_summary['A4'], ws_summary['B4']]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

                df = st.session_state.df_analysis
                row = 5
                summary_data = [
                    ('총 SKU 수', len(df)),
                    ('재고 위험 품목 (≤7일)', len(df[df['재고소진일'] <= 7])),
                    ('발주 필요 품목', len(df[df['발주필요'] == True])),
                    ('과다 재고 품목', len(df[df['재고상태'] == '🔵 과잉'])),
                    ('평균 재고소진일', f"{df[df['재고소진일'] < 999]['재고소진일'].mean():.1f}일"),
                ]
                for label, value in summary_data:
                    ws_summary.cell(row, 1, label)
                    ws_summary.cell(row, 2, value)
                    row += 1

                # 시트 2: 긴급 발주 품목
                ws_urgent = wb.create_sheet("긴급발주품목")
                urgent_items = df[df['재고소진일'] <= 7].sort_values('재고소진일')

                headers = ['SKU코드', '제품명', 'ABC', 'XYZ', '카테고리', '현재고', '안전재고', '재고소진일', '권장발주량']
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws_urgent.cell(1, col_idx, header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")

                for row_idx, (_, item) in enumerate(urgent_items.iterrows(), start=2):
                    ws_urgent.cell(row_idx, 1, item['SKU코드'])
                    ws_urgent.cell(row_idx, 2, item['제품명'])
                    ws_urgent.cell(row_idx, 3, item.get('ABC등급', ''))
                    ws_urgent.cell(row_idx, 4, item.get('XYZ등급', ''))
                    ws_urgent.cell(row_idx, 5, item.get('카테고리', ''))
                    ws_urgent.cell(row_idx, 6, item['현재고'])
                    ws_urgent.cell(row_idx, 7, item['안전재고'])
                    ws_urgent.cell(row_idx, 8, item['재고소진일'])
                    ws_urgent.cell(row_idx, 9, item.get('권장발주량', 0))

                # 시트 3: 발주 필요 품목
                ws_order = wb.create_sheet("발주필요품목")
                order_items = df[(df['발주필요'] == True) & (df['권장발주량'] > 0)].sort_values('현재고')

                for col_idx, header in enumerate(headers, start=1):
                    cell = ws_order.cell(1, col_idx, header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFF4E5", end_color="FFF4E5", fill_type="solid")

                for row_idx, (_, item) in enumerate(order_items.iterrows(), start=2):
                    ws_order.cell(row_idx, 1, item['SKU코드'])
                    ws_order.cell(row_idx, 2, item['제품명'])
                    ws_order.cell(row_idx, 3, item.get('ABC등급', ''))
                    ws_order.cell(row_idx, 4, item.get('XYZ등급', ''))
                    ws_order.cell(row_idx, 5, item.get('카테고리', ''))
                    ws_order.cell(row_idx, 6, item['현재고'])
                    ws_order.cell(row_idx, 7, item['안전재고'])
                    ws_order.cell(row_idx, 8, item['재고소진일'])
                    ws_order.cell(row_idx, 9, item.get('권장발주량', 0))

                # 컬럼 너비 자동 조정
                for ws in [ws_summary, ws_urgent, ws_order]:
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

                # 바이트로 저장
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                st.download_button(
                    label="📥 일일 리포트 다운로드",
                    data=buffer,
                    file_name=f"일일리포트_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    key="daily_report_download"
                )
            else:
                st.warning("⚠️ 데이터를 먼저 로드하세요")

    with col_info_report:
        st.info("""
        **📋 일일 리포트 내용**

        - **일일요약**: 전체 재고 현황 요약
        - **긴급발주품목**: 재고소진일 ≤ 7일인 위험 품목
        - **발주필요품목**: 발주점 이하 전체 품목 리스트
        - **카테고리 정보 포함**
        """)

    st.markdown("---")

    if st.button("💾 설정 저장", type="primary"):
        st.success("✅ 설정이 저장되었습니다!")

# 앱 실행
if __name__ == "__main__":
    main()
